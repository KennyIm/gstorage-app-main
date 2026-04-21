from django.shortcuts import render
import os
from django.conf import settings
from django.urls import reverse_lazy
from django.db import transaction
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from rest_framework.pagination import PageNumberPagination
from rest_framework import generics, permissions
from django.db.models import Count
from django.db.models import Sum
from django.db.models.functions import TruncMonth
from datetime import datetime, timedelta
from django.utils import timezone
from rest_framework.views import APIView     
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
from .utils import actualizar_estados_automaticos, registrar_auditoria, generar_numeros_orden_despacho
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.worksheet.page import PageMargins
from django.http import HttpResponse
from rest_framework import status
from django.shortcuts import get_object_or_404
from rest_framework.exceptions import PermissionDenied
from django.db.models import Q, Exists, OuterRef
from django.views.generic import (
    ListView, 
    DetailView, 
    CreateView, 
    UpdateView, 
    DeleteView
)
from .models import (
    Mercancia, Cliente, Despacho, Conductor, 
    Camion, Ruta, Destino, Ubicacion, HistorialMovimientos, Estanteria,
    AreaRestringida, Proveedor, Rampla, Cotizacion, PermisoColaboracion
)

from .serializers import (
    MercanciaListSerializer, MercanciaWriteSerializer,
    DespachoListSerializer, DespachoWriteSerializer,
    ClienteSerializer, ConductorSerializer, CamionSerializer, 
    RutaSerializer, DestinoSerializer, UbicacionSerializer, EstanteriaSerializer,
    HistorialSerializer,
    AreaRestringidaSerializer,
    ProveedorSerializer,
    RamplaSerializer,
    CotizacionSerializer,
    InvitacionColaboradorSerializer,
)
from usuarios.permissions import IsAdminEmpresa, IsJefeDeBodega, IsOperario

from django.contrib.auth.models import User

from django import forms

def get_empresa_from_user(request):
    if not request.user.is_authenticated or not hasattr(request.user, 'perfil'):
        raise ValidationError("Usuario no autenticado o sin perfil.")
    empresa = request.user.perfil.empresa
    if not empresa:
        raise ValidationError("El usuario no está asociado a ninguna empresa.")
    return empresa

def filtrar_por_sucursal_y_empresa(queryset, request):
    user = request.user
    if not hasattr(user, 'perfil') or not user.perfil.empresa:
        return queryset.none()
        
    perfil = user.perfil
    qs = queryset.filter(empresa=perfil.empresa)
    sucursal_id_solicitada = request.query_params.get('sucursal_id')

    if perfil.rol == 'DUENO':
        if sucursal_id_solicitada:
            qs = qs.filter(sucursal_id=sucursal_id_solicitada)
    else:
        if sucursal_id_solicitada:
            qs = qs.filter(sucursal_id=sucursal_id_solicitada)
        else:
            if perfil.sucursal:
                qs = qs.filter(sucursal=perfil.sucursal)
            else:
                return queryset.none()
            
    return qs
# --- Vistas de API para Mercancia ---

class MercanciaListCreateAPI(generics.ListCreateAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        empresa = get_empresa_from_user(self.request)
        qs = Mercancia.activos.filter(empresa=empresa)
        
        despacho_id = self.request.query_params.get('id_despacho')
        estado = self.request.query_params.get('estado')
        estado_in = self.request.query_params.get('estado_in')

        if user.perfil.rol != 'DUENO':
            sucursal_empleado = user.perfil.sucursal
            condicion_propia = Q(sucursal=sucursal_empleado)
            
            if despacho_id:
                condicion_invitado = Q(
                    id_despacho__colaboradores_invitados__usuario_invitado=user,
                    id_despacho__colaboradores_invitados__activo=True
                )
                qs = qs.filter(condicion_propia | condicion_invitado).distinct()
            else:
                qs = qs.filter(condicion_propia)

        if despacho_id:
            qs = qs.filter(id_despacho=despacho_id)
        if estado:
            qs = qs.filter(estado=estado)
        if estado_in:
            estados = estado_in.split(',')
            qs = qs.filter(estado__in=estados)

        return qs.order_by('-fecha_ingreso')
    

    def get_serializer_class(self):
        if self.request.method == 'GET':
            return MercanciaListSerializer
        return MercanciaWriteSerializer
    
    def perform_create(self, serializer):
        empresa = get_empresa_from_user(self.request)
        user = self.request.user
        sucursal_empleado = user.perfil.sucursal if hasattr(user, 'perfil') else None
        
        nuevos_datos = serializer.validated_data
        ubicacion_seleccionada = nuevos_datos.get('id_ubicacion_actual')
        nuevo_kg = nuevos_datos.get('kg')
        nuevo_m3 = nuevos_datos.get('m3')

        # --- VALIDACIONES DE CREACIÓN ---
        if ubicacion_seleccionada:

            if ubicacion_seleccionada.empresa != empresa:
                raise ValidationError({"id_ubicacion_actual": "Esta ubicación no pertenece a su empresa."})
            
            if ubicacion_seleccionada.estado_ocupado:
                raise ValidationError(
                    {"id_ubicacion_actual": f"¡Conflicto! La ubicación {ubicacion_seleccionada.codigo_ubicacion} ya está ocupada por otro lote."}
                )

            if nuevo_kg and ubicacion_seleccionada.capacidad_maxima_kg:
                if float(nuevo_kg) > float(ubicacion_seleccionada.capacidad_maxima_kg):
                    raise ValidationError({
                        "id_ubicacion_actual": f"Exceso de Peso: El lote ({nuevo_kg}kg) supera la capacidad de la ubicación ({ubicacion_seleccionada.capacidad_maxima_kg}kg)."
                    })

            if nuevo_m3 and ubicacion_seleccionada.capacidad_max_m3:
                if float(nuevo_m3) > float(ubicacion_seleccionada.capacidad_max_m3):
                    raise ValidationError({
                        "id_ubicacion_actual": f"Exceso de Volumen: El lote ({nuevo_m3}m³) supera la capacidad de la ubicación ({ubicacion_seleccionada.capacidad_max_m3}m³)."
                    })
        id_despacho = nuevos_datos.get('id_despacho')
        sucursal_final = id_despacho.sucursal if id_despacho else sucursal_empleado
        instance = serializer.save(
            id_usuario_creacion=user,
            estado='En Bodega',
            activo=True,
            empresa=empresa,
            sucursal=sucursal_final
        )
        try:
            if instance.id_ubicacion_actual:
                instance.id_ubicacion_actual.estado_ocupado = True
                instance.id_ubicacion_actual.save()
        except Exception as e:
            print(f"Error al ocupar ubicación: {e}")
        HistorialMovimientos.objects.create(
            empresa=empresa,
            sucursal=sucursal_empleado,
            id_mercancia=instance,
            id_usuario=user,
            id_ubicacion_anterior=None,
            id_ubicacion_nueva=instance.id_ubicacion_actual,
            accion='Creación',
            descripcion_adicional=f"Mercancía creada en {instance.id_ubicacion_actual.codigo_ubicacion if instance.id_ubicacion_actual else 'Bodega General'}"
        )
    

class MercanciaDetailAPI(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = MercanciaWriteSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        empresa = get_empresa_from_user(self.request)
        
        if not hasattr(user, 'perfil'):
            return Mercancia.objects.none()

        qs = Mercancia.objects.filter(empresa=empresa)

        if user.perfil.rol == 'DUENO':
            return qs

        sucursal_empleado = user.perfil.sucursal

        condicion_sucursal = Q(sucursal=sucursal_empleado)

        condicion_invitacion = Q(
            id_despacho__colaboradores_invitados__usuario_invitado=user,
            id_despacho__colaboradores_invitados__activo=True
        )

        return qs.filter(condicion_sucursal | condicion_invitacion).distinct()

    def perform_update(self, serializer):
        empresa = get_empresa_from_user(self.request)
        user = self.request.user if self.request.user.is_authenticated else None
        
        sucursal_empleado = user.perfil.sucursal if hasattr(user, 'perfil') else None

        instance = self.get_object() 
        ubicacion_original = instance.id_ubicacion_actual
        estado_original = instance.estado

        datos_nuevos = serializer.validated_data
        
        nueva_ubicacion = datos_nuevos.get('id_ubicacion_actual', instance.id_ubicacion_actual)
        nuevo_estado = datos_nuevos.get('estado', instance.estado)
        nuevo_kg = datos_nuevos.get('kg', instance.kg)
        nuevo_m3 = datos_nuevos.get('m3', instance.m3)

        if nueva_ubicacion and (nueva_ubicacion != ubicacion_original or nuevo_estado in ['En Bodega', 'Asignado']):
            if nueva_ubicacion.estado_ocupado:
                ocupante = Mercancia.activos.filter(id_ubicacion_actual=nueva_ubicacion).exclude(pk=instance.pk).first()
                if ocupante:
                    raise ValidationError({
                        "id_ubicacion_actual": f"¡Conflicto! La ubicación {nueva_ubicacion.codigo_ubicacion} ya está ocupada por el Lote #{ocupante.id_mercancia}."
                    })
            
            if nueva_ubicacion.empresa != empresa:
                raise ValidationError({"id_ubicacion_actual": "Esta ubicación no pertenece a su empresa."})

            if nuevo_kg and nueva_ubicacion.capacidad_maxima_kg:
                if float(nuevo_kg) > float(nueva_ubicacion.capacidad_maxima_kg):
                    raise ValidationError({
                        "id_ubicacion_actual": f"Exceso de Peso: El lote ({nuevo_kg}kg) supera la capacidad de la ubicación ({nueva_ubicacion.capacidad_maxima_kg}kg)."
                    })

            if nuevo_m3 and nueva_ubicacion.capacidad_max_m3:
                if float(nuevo_m3) > float(nueva_ubicacion.capacidad_max_m3):
                    raise ValidationError({
                        "id_ubicacion_actual": f"Exceso de Volumen: El lote ({nuevo_m3}m³) supera la capacidad de la ubicación ({nueva_ubicacion.capacidad_max_m3}m³)."
                    })

        instance_actualizada = serializer.save(id_usuario_ultima_modificacion=user)

        log_especial_creado = False

        if nuevo_estado in ['Merma', 'Eliminado']:
            if ubicacion_original:
                try:
                    ubicacion_original.estado_ocupado = False
                    ubicacion_original.save()
                except Exception as e:
                    print(f"Error liberando: {e}")
            
            instance_actualizada.id_despacho = None
            instance_actualizada.id_ubicacion_actual = None
            instance_actualizada.save()
            
            HistorialMovimientos.objects.create(
                empresa=empresa, 
                sucursal=sucursal_empleado, 
                id_mercancia=instance_actualizada, 
                id_usuario=user,
                id_ubicacion_anterior=ubicacion_original, 
                id_ubicacion_nueva=None,
                modelo_afectado="Mercancía", 
                accion='Edición (Baja)', 
                descripcion_adicional=f"Mercancía marcada como {nuevo_estado}"
            )
            return

        if ubicacion_original != nueva_ubicacion:
            if ubicacion_original:
                ubicacion_original.estado_ocupado = False
                ubicacion_original.save()
            
            if nueva_ubicacion and nuevo_estado in ['En Bodega', 'Asignado']:
                nueva_ubicacion.estado_ocupado = True
                nueva_ubicacion.save()
            
            HistorialMovimientos.objects.create(
                empresa=empresa, 
                sucursal=sucursal_empleado, 
                id_mercancia=instance_actualizada, 
                id_usuario=user,
                id_ubicacion_anterior=ubicacion_original, 
                id_ubicacion_nueva=nueva_ubicacion,
                modelo_afectado="Mercancía", 
                accion='Reubicación', 
                descripcion_adicional=f"Movido de {ubicacion_original.codigo_ubicacion if ubicacion_original else 'Nada'} a {nueva_ubicacion.codigo_ubicacion if nueva_ubicacion else 'Nada'}"
            )
            log_especial_creado = True

        elif estado_original != nuevo_estado:
            if nuevo_estado in ['En Tránsito', 'Entregado']:
                if ubicacion_original:
                    ubicacion_original.estado_ocupado = False
                    ubicacion_original.save()
                instance_actualizada.id_ubicacion_actual = None
                instance_actualizada.save()

            elif nuevo_estado in ['En Bodega', 'Asignado']:
                if ubicacion_original:
                    ubicacion_original.estado_ocupado = True
                    ubicacion_original.save()

            HistorialMovimientos.objects.create(
                empresa=empresa, 
                sucursal=sucursal_empleado, 
                id_mercancia=instance_actualizada, 
                id_usuario=user,
                id_ubicacion_anterior=ubicacion_original, 
                id_ubicacion_nueva=instance_actualizada.id_ubicacion_actual,
                modelo_afectado="Mercancía", 
                accion='Cambio de Estado',
                descripcion_adicional=f"Cambio de estado de '{estado_original}' a '{nuevo_estado}'"
            )
            log_especial_creado = True

        if not log_especial_creado:
            HistorialMovimientos.objects.create(
                empresa=empresa, 
                sucursal=instance_actualizada.sucursal,
                id_mercancia=instance_actualizada, 
                id_usuario=user,
                id_ubicacion_anterior=ubicacion_original, 
                id_ubicacion_nueva=instance_actualizada.id_ubicacion_actual,
                modelo_afectado="Mercancía",
                accion='Edición General',
                descripcion_adicional=f"Se actualizaron datos generales de la mercancía con el código: {instance_actualizada.codigo_interno or 'Sin código'}, pertenece al cliente: {instance_actualizada.id_cliente or 'N/N'}. N° Factura: {instance_actualizada.factura or 'N/N'}"
            )

    def perform_destroy(self, instance):
        empresa = get_empresa_from_user(self.request)
        user = self.request.user if self.request.user.is_authenticated else None
        sucursal_empleado = user.perfil.sucursal if hasattr(user, 'perfil') else None

        HistorialMovimientos.objects.create(
            empresa=empresa, 
            sucursal=sucursal_empleado,
            id_mercancia=instance, 
            id_usuario=user,
            id_ubicacion_anterior=instance.id_ubicacion_actual, 
            id_ubicacion_nueva=None,
            accion='Borrado Lógico',
            modelo_afectado="Mercancía", 
            descripcion_adicional=f"Mercancía eliminada (Motivo: {instance.motivo_baja or 'No especificado'})"
        )
        
        if instance.id_ubicacion_actual:
            try:
                ubicacion = instance.id_ubicacion_actual
                ubicacion.estado_ocupado = False
                ubicacion.save()
            except Exception as e:
                print(f"Error al liberar ubicación en destroy: {e}")

        instance.id_despacho = None
        instance.id_ubicacion_actual = None
        instance.activo = False
        instance.estado = 'Eliminado'
        instance.id_usuario_ultima_modificacion = user
        instance.save()

class MercanciaAsignarMasivoAPI(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        ids = request.data.get('ids', [])
        id_despacho = request.data.get('id_despacho')
        user = request.user
        if not ids or not id_despacho:
            return Response(
                {"error": "Debe seleccionar mercancías y un despacho de destino."}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            with transaction.atomic():
                queryset_seguro = filtrar_por_sucursal_y_empresa(Mercancia.objects.all(), request)
                mercancias = queryset_seguro.filter(id_mercancia__in=ids)

                if not mercancias.exists():
                    return Response(
                        {"error": "No se encontraron mercancías válidas para procesar en su sucursal."}, 
                        status=status.HTTP_404_NOT_FOUND
                    )
                ubicaciones_ids = mercancias.exclude(id_ubicacion_actual__isnull=True)\
                                            .values_list('id_ubicacion_actual', flat=True)

                count = mercancias.update(
                    id_despacho_id=id_despacho,
                    id_ubicacion_actual=None,
                    estado='Asignado',
                    id_usuario_ultima_modificacion=user
                )

                if ubicaciones_ids:
                    Ubicacion.objects.filter(id_ubicacion__in=ubicaciones_ids).update(estado_ocupado=False)

                empresa = get_empresa_from_user(request)
                HistorialMovimientos.objects.create(
                    empresa=empresa,
                    id_usuario=user,
                    id_mercancia=None,
                    tipo_movimiento='Asignación Masiva',
                    descripcion_adicional=f"Asignación masiva de {mercancias.count()} ítems al Despacho #{id_despacho}."
                    )

            return Response({
                "message": f"Éxito: {count} mercancías asignadas y ubicaciones liberadas."
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": f"Error en el servidor: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
# --- Vistas de API para Despacho ---

class DespachoListCreateAPI(generics.ListCreateAPIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        empresa = get_empresa_from_user(self.request)
        perfil = user.perfil
        
        actualizar_estados_automaticos(empresa)

        es_colaborador_subquery = PermisoColaboracion.objects.filter(
            despacho_id=OuterRef('pk'),
            usuario_invitado=user,
            activo=True
        )

        qs = Despacho.objects.filter(empresa=empresa, activo=True).annotate(
            es_colaborador=Exists(es_colaborador_subquery)
        )

        sucursal_id_solicitada = self.request.query_params.get('sucursal_id')

        if perfil.rol == 'DUENO':
            if sucursal_id_solicitada and sucursal_id_solicitada != 'null':
                qs = qs.filter(sucursal_id=sucursal_id_solicitada)
            return qs.order_by('-fecha_programada')

        condicion_invitado = Q(es_colaborador=True)

        if sucursal_id_solicitada and sucursal_id_solicitada != 'null':
            condicion_sucursal = Q(sucursal_id=sucursal_id_solicitada)
            return qs.filter(condicion_sucursal | condicion_invitado).distinct().order_by('-fecha_programada')
        else:
            condicion_propio = Q(sucursal=perfil.sucursal)
            return qs.filter(condicion_propio | condicion_invitado).distinct().order_by('-fecha_programada')
    
    def get_serializer_class(self):
        if self.request.method == 'GET':
            return DespachoListSerializer
        return DespachoWriteSerializer
    
    def perform_create(self, serializer):
        empresa = get_empresa_from_user(self.request)
        user = self.request.user
        sucursal_empleado = user.perfil.sucursal if hasattr(user, 'perfil') else None
        
        instance = serializer.save(
            id_usuario_creacion=user, 
            empresa=empresa,
            sucursal=sucursal_empleado, 
            activo=True
        )
        registrar_auditoria(
            empresa=empresa,
            usuario=user,
            modelo="Despacho",
            sucursal=self.request.user.perfil.sucursal,
            accion="Creación",
            descripcion=f"Despacho programado para {instance.fecha_programada}"
        )

class DespachoDetailAPI(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DespachoWriteSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        #empresa = get_empresa_from_user(self.request)
        #return Despacho.objects.filter(empresa=empresa, activo=True)
        empresa = get_empresa_from_user(self.request)
        return Despacho.objects.filter(empresa=empresa, activo=True)
    
    def perform_update(self, serializer):
        instance = self.get_object()
        user_perfil = self.request.user.perfil

        if user_perfil.rol != 'DUENO' and instance.sucursal_id != user_perfil.sucursal_id:
            raise PermissionDenied("No tienes permisos para modificar despachos de otras sucursales.")

        prev_estado = instance.estado_despacho
        instance = serializer.save(id_usuario_ultima_modificacion=self.request.user)
        
        desc = f"Despacho #{instance.id_despacho} actualizado."
        if prev_estado != instance.estado_despacho:
            desc += f" Estado cambió de {prev_estado} a {instance.estado_despacho}."

        registrar_auditoria(
            empresa=instance.empresa,
            usuario=self.request.user,
            modelo="Despacho",
            accion="Edición",
            sucursal=self.request.user.perfil.sucursal,
            descripcion=desc
        )

    def perform_destroy(self, instance):
        user_perfil = self.request.user.perfil

        if instance.estado_despacho == 'Finalizado':
            raise ValidationError({
                "error": "No se puede eliminar un despacho Finalizado. La mercancía ya fue entregada y borrarlo causaría inconsistencias graves en el inventario."
            })

        if user_perfil.rol != 'DUENO' and instance.sucursal_id != user_perfil.sucursal_id:
            raise PermissionDenied("No tienes permisos para eliminar despachos de otras sucursales.")

        instance.activo = False
        instance.estado_despacho = 'Eliminado'
        instance.id_usuario_ultima_modificacion = self.request.user
        instance.save()
        
        Mercancia.objects.filter(id_despacho=instance).update(
            id_despacho=None,
            estado='En Bodega'
        )
        
        if instance.id_camion:
            instance.id_camion.estado_camion = 'DISPONIBLE'
            instance.id_camion.save()
            
        if hasattr(instance, 'id_rampla') and instance.id_rampla:
            instance.id_rampla.estado_rampla = 'Disponible'
            instance.id_rampla.save()

        registrar_auditoria(
            empresa=instance.empresa,
            usuario=self.request.user,
            sucursal=self.request.user.perfil.sucursal,
            modelo="Despacho",
            accion="Eliminación",
            descripcion=f"Se eliminó (lógico) el despacho: {instance.id_despacho}. Se retornó su carga a bodega y se liberaron los vehículos."
        )

# --- Vistas de API para Clientes ---

class ClienteListCreateAPI(generics.ListCreateAPIView):
    serializer_class = ClienteSerializer
    permission_classes = [permissions.IsAuthenticated, IsJefeDeBodega]

    def get_queryset(self):
        empresa = get_empresa_from_user(self.request)
        return Cliente.objects.filter(empresa=empresa)
    
    def perform_create(self, serializer):
        empresa = get_empresa_from_user(self.request)
        instance=serializer.save(empresa=empresa, activo=True)
        registrar_auditoria(
            empresa=empresa,
            usuario=self.request.user,
            sucursal=self.request.user.perfil.sucursal,
            modelo="Cliente",
            accion="Creación",
            descripcion=f"Se creó el cliente: {instance.nombre_cliente}"
        )

class ClienteDetailAPI(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = ClienteSerializer
    permission_classes = [permissions.IsAuthenticated, IsJefeDeBodega]
    
    def get_queryset(self):
        empresa = get_empresa_from_user(self.request)
        return Cliente.objects.filter(empresa=empresa)
    
    def perform_update(self, serializer):
        instance = serializer.save()
        registrar_auditoria(
            empresa=instance.empresa,
            usuario=self.request.user,
            sucursal=self.request.user.perfil.sucursal,
            modelo="Cliente",
            accion="Edición",
            descripcion=f"Se actualizaron datos del cliente: {instance.nombre_cliente}"
        )
    
    def perform_destroy(self, instance):
        instance.activo = False
        instance.save()
        registrar_auditoria(
            empresa=instance.empresa,
            usuario=self.request.user,
            sucursal=self.request.user.perfil.sucursal,
            modelo="Cliente",
            accion="Eliminación",
            descripcion=f"Se eliminó (lógico) al cliente: {instance.nombre_cliente}"
        )

# --- Vistas de API para Conductores ---

class ConductorListCreateAPI(generics.ListCreateAPIView):
    serializer_class = ConductorSerializer
    permission_classes = [permissions.IsAuthenticated, IsJefeDeBodega]
    
    def get_queryset(self):
        empresa = get_empresa_from_user(self.request)
        return Conductor.objects.filter(empresa=empresa)
    
    def perform_create(self, serializer):
        empresa = get_empresa_from_user(self.request)
        instance = serializer.save(empresa=empresa, activo=True)
        registrar_auditoria(
            empresa=empresa,
            usuario=self.request.user,
            sucursal=self.request.user.perfil.sucursal,
            modelo="Conductor",
            accion="Creación",
            descripcion=f"Se creó el conductor: {instance.nombre_completo}"
        )

class ConductorDetailAPI(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = ConductorSerializer
    permission_classes = [permissions.IsAuthenticated, IsJefeDeBodega]

    def get_queryset(self):
        empresa = get_empresa_from_user(self.request)
        return Conductor.objects.filter(empresa=empresa)
    
    def perform_update(self, serializer):
        instance = serializer.save()
        registrar_auditoria(
            empresa=instance.empresa,
            usuario=self.request.user,
            sucursal=self.request.user.perfil.sucursal,
            modelo="Conductor",
            accion="Edición",
            descripcion=f"Se actualizarón datos del conductor: {instance.nombre_completo}"
        )
    
    def perform_destroy(self, instance): 
        instance.activo = False
        instance.save()
        registrar_auditoria(
            empresa=instance.empresa,
            usuario=self.request.user,
            sucursal=self.request.user.perfil.sucursal,
            modelo="Conductor",
            accion="Eliminación",
            descripcion=f"Se eliminó al conductor: {instance.nombre_completo}"
        )

# --- Vistas de API para Camiones ---

class CamionListCreateAPI(generics.ListCreateAPIView):
    serializer_class = CamionSerializer 
    permission_classes = [permissions.IsAuthenticated, IsJefeDeBodega]
    
    def get_queryset(self):
        empresa = get_empresa_from_user(self.request)
        return Camion.objects.filter(empresa=empresa)
    
    def perform_create(self, serializer):
        empresa = get_empresa_from_user(self.request)
        instance = serializer.save(empresa=empresa, activo=True)
        registrar_auditoria(
            empresa=empresa,
            usuario=self.request.user,
            sucursal=self.request.user.perfil.sucursal,
            modelo="Camion",
            accion="Creación",
            descripcion=f"Se creó el camión: {instance.patente}"
        )

class CamionDetailAPI(generics.RetrieveUpdateDestroyAPIView):
    queryset = Camion.objects.all() 
    serializer_class = CamionSerializer
    permission_classes = [permissions.IsAuthenticated, IsJefeDeBodega]

    def get_queryset(self):
        empresa = get_empresa_from_user(self.request)
        return Camion.objects.filter(empresa=empresa)
    
    def perform_update(self, serializer):
        instance = serializer.save()
        registrar_auditoria(
            empresa=instance.empresa,
            usuario=self.request.user,
            sucursal=self.request.user.perfil.sucursal,
            modelo="Camion",
            accion="Edición",
            descripcion=f"Se actualizarón datos del camión: {instance.patente}"
        )
    
    def perform_destroy(self, instance):
        instance.activo = False
        instance.save()
        registrar_auditoria(
            empresa=instance.empresa,
            usuario=self.request.user,
            sucursal=self.request.user.perfil.sucursal,
            modelo="Camion",
            accion="Eliminación",
            descripcion=f"Se eliminó el camión: {instance.patente}"
        )

# --- Vistas de API para Rutas ---

class RutaListCreateAPI(generics.ListCreateAPIView):
    serializer_class = RutaSerializer
    permission_classes = [permissions.IsAuthenticated, IsJefeDeBodega]
    
    def get_queryset(self):
        #empresa = get_empresa_from_user(self.request)
        #return Ruta.objects.filter(empresa=empresa)
        return filtrar_por_sucursal_y_empresa(Ruta.objects.all(), self.request)
    
    def perform_create(self, serializer):
        empresa = get_empresa_from_user(self.request)
        user = self.request.user
        sucursal_empleado = user.perfil.sucursal if hasattr(user, 'perfil') else None
        instance = serializer.save(
            empresa=empresa, 
            sucursal=sucursal_empleado, 
            activo=True
        )
        registrar_auditoria(
            empresa=empresa,
            usuario=self.request.user,
            sucursal=self.request.user.perfil.sucursal,
            modelo="Ruta",
            accion="Creación",
            descripcion=f"Se creó la ruta: {instance.nombre_ruta}"
        )

class RutaDetailAPI(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = RutaSerializer
    permission_classes = [permissions.IsAuthenticated, IsJefeDeBodega]

    def get_queryset(self):
        #empresa = get_empresa_from_user(self.request)
        #return Ruta.objects.filter(empresa=empresa)
        return filtrar_por_sucursal_y_empresa(Ruta.objects.all(), self.request)
    
    def perform_update(self, serializer):
        instance = serializer.save()
        registrar_auditoria(
            empresa=instance.empresa,
            usuario=self.request.user,
            sucursal=self.request.user.perfil.sucursal,
            modelo="Ruta",
            accion="Edición",
            descripcion=f"Se actualizarón datos de la ruta: {instance.nombre_ruta}"
        )
    
    def perform_destroy(self, instance):
        instance.activo = False
        instance.save()
        registrar_auditoria(
            empresa=instance.empresa,
            usuario=self.request.user,
            sucursal=self.request.user.perfil.sucursal,
            modelo="Ruta",
            accion="Eliminación",
            descripcion=f"Se eliminó la ruta: {instance.nombre_ruta}"
        )

# --- Vistas de API para Destinos ---

class DestinoListCreateAPI(generics.ListCreateAPIView):
    serializer_class = DestinoSerializer
    permission_classes = [permissions.IsAuthenticated, IsJefeDeBodega]

    def get_queryset(self):
        empresa = get_empresa_from_user(self.request)
        return Destino.objects.filter(empresa=empresa)
    
    def perform_create(self, serializer):
        empresa = get_empresa_from_user(self.request)
        instance = serializer.save(empresa=empresa, activo=True)
        registrar_auditoria(
            empresa=empresa,
            usuario=self.request.user,
            sucursal=self.request.user.perfil.sucursal,
            modelo="Destino",
            accion="Creación",
            descripcion=f"Se creó el destino: {instance.nombre_ciudad}"
        )

class DestinoDetailAPI(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DestinoSerializer
    permission_classes = [permissions.IsAuthenticated, IsJefeDeBodega]

    def get_queryset(self):
        empresa = get_empresa_from_user(self.request)
        return Destino.objects.filter(empresa=empresa)
    
    def perform_update(self, serializer):
        instance = serializer.save()
        registrar_auditoria(
            empresa=instance.empresa,
            usuario=self.request.user,
            sucursal=self.request.user.perfil.sucursal,
            modelo="Destino",
            accion="Edición",
            descripcion=f"Se actualizarón datos del destino: {instance.nombre_ciudad}"
        )
    
    def perform_destroy(self, instance):
        instance.activo = False
        instance.save()
        registrar_auditoria(
            empresa=instance.empresa,
            usuario=self.request.user,
            sucursal=self.request.user.perfil.sucursal,
            modelo="Destino",
            accion="Eliminación",
            descripcion=f"Se eliminó el destino: {instance.nombre_ciudad}"
        )

# --- Vistas de API para Ubicaciones ---

class UbicacionListCreateAPI(generics.ListCreateAPIView):
    serializer_class = UbicacionSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return filtrar_por_sucursal_y_empresa(Ubicacion.activos.all(), self.request)
    
    def perform_create(self, serializer):
        empresa = get_empresa_from_user(self.request)
        user = self.request.user
        sucursal_empleado = user.perfil.sucursal if hasattr(user, 'perfil') else None 
        
        serializer.save(empresa=empresa, sucursal=sucursal_empleado, activo=True)

class UbicacionDetailAPI(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = UbicacionSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return filtrar_por_sucursal_y_empresa(Ubicacion.activos.all(), self.request)
    
    def perform_destroy(self, instance):
        if instance.estado_ocupado:
            mercancia = Mercancia.activos.filter(id_ubicacion_actual=instance).exists()
            if mercancia:
                raise ValidationError(
                    {"detail": f"No se puede eliminar la ubicación {instance.codigo_ubicacion}: Tiene mercancía asignada."}
                )

        instance.activo = False
        instance.save()


class DashboardStatsAPI(APIView):
    permission_classes = [permissions.IsAuthenticated, IsJefeDeBodega]

    def get(self, request, format=None):
        empresa = get_empresa_from_user(request)
        
        total_en_bodega = Mercancia.activos.filter(empresa=empresa, estado='En Bodega').count()
        despachos_programados = Despacho.objects.filter(empresa=empresa, estado_despacho='Programado').count()
        ubicaciones_libres = Ubicacion.activos.filter(empresa=empresa, estado_ocupado=False).count()
        total_clientes = Cliente.activos.filter(empresa=empresa).count()

        context = {
            'total_en_bodega': total_en_bodega,
            'despachos_programados': despachos_programados,
            'ubicaciones_libres': ubicaciones_libres,
            'total_clientes': total_clientes,
        }
        
        return Response(context)
    
class EstanteriaListCreateAPI(generics.ListCreateAPIView):
    serializer_class = EstanteriaSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return filtrar_por_sucursal_y_empresa(Estanteria.objects.all(), self.request)
    
    def perform_create(self, serializer):
        empresa = get_empresa_from_user(self.request)
        user = self.request.user
        sucursal_empleado = user.perfil.sucursal if hasattr(user, 'perfil') else None
        
        serializer.save(empresa=empresa, sucursal=sucursal_empleado, activo=True)

class EstanteriaDetailAPI(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = EstanteriaSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return filtrar_por_sucursal_y_empresa(Estanteria.objects.all(), self.request)

    def perform_update(self, serializer):
        instance = self.get_object()
        old_ancho = instance.num_modulos_ancho
        old_alto = instance.num_niveles_alto
        old_prof = instance.num_profundidad
        
        new_ancho = serializer.validated_data.get('num_modulos_ancho', old_ancho)
        new_alto = serializer.validated_data.get('num_niveles_alto', old_alto)
        new_prof = serializer.validated_data.get('num_profundidad', old_prof)

        if new_ancho < old_ancho or new_alto < old_alto or new_prof < old_prof:
            ubicaciones_afectadas = Ubicacion.activos.filter(
                estanteria=instance
            ).filter(
                Q(pos_x_rel__gte=new_ancho) |     
                Q(pos_y_rel__gt=new_alto) |        
                Q(pos_z_rel__gte=new_prof)         
            )

            if ubicaciones_afectadas.filter(estado_ocupado=True).exists():
                raise ValidationError(
                    {"detail": f"No se puede reducir el tamaño: Hay mercancía en las ubicaciones que intentas eliminar. Mueve la carga primero."}
                )
            ids_a_borrar = list(ubicaciones_afectadas.values_list('id_ubicacion', flat=True))
        else:
            ids_a_borrar = []

        estanteria = serializer.save()

        if ids_a_borrar:
            Ubicacion.objects.filter(id_ubicacion__in=ids_a_borrar).delete() 

        if new_ancho > old_ancho or new_alto > old_alto or new_prof > old_prof:
             self.rellenar_huecos(estanteria)

    def perform_destroy(self, instance):
        ubicaciones_ocupadas = Ubicacion.activos.filter(estanteria=instance, estado_ocupado=True)
        if ubicaciones_ocupadas.exists():
            count = ubicaciones_ocupadas.count()
            raise ValidationError(
                {"detail": f"No se puede eliminar la estantería: Contiene {count} ubicaciones con mercancía. Vacíela primero."}
            )
        Ubicacion.activos.filter(estanteria=instance).update(activo=False)
        instance.activo = False
        instance.save()

    def rellenar_huecos(self, estanteria):
        total_niveles = estanteria.num_niveles_alto + 1
        volumen = estanteria.ancho_hueco_m * estanteria.alto_hueco_m * estanteria.profundo_hueco_m
        capacidad = estanteria.capacidad_carga_por_hueco_kg

        ubicaciones_crear = []
        for x in range(estanteria.num_modulos_ancho):
            for y in range(total_niveles):
                for z in range(estanteria.num_profundidad):
                    if not Ubicacion.objects.filter(estanteria=estanteria, pos_x_rel=x, pos_y_rel=y, pos_z_rel=z).exists():
                        codigo = f"{estanteria.codigo_estanteria}-M{x}-N{y}-P{z}"
                        
                        ubicaciones_crear.append(Ubicacion(
                            empresa=estanteria.empresa,
                            sucursal=estanteria.sucursal,
                            estanteria=estanteria,
                            codigo_ubicacion=codigo,
                            pos_x_rel=x, pos_y_rel=y, pos_z_rel=z,
                            capacidad_maxima_kg=capacidad,
                            capacidad_max_m3=volumen,
                            es_zona_suelo=False,
                            estado_ocupado=False
                        ))
        
        if ubicaciones_crear:
            Ubicacion.objects.bulk_create(ubicaciones_crear)
    

class HistorialPagination(PageNumberPagination):
    page_size = 25  
    page_size_query_param = 'page_size'
    max_page_size = 50

class HistorialListAPI(generics.ListAPIView):
    serializer_class = HistorialSerializer
    permission_classes = [permissions.IsAuthenticated, IsJefeDeBodega]
    pagination_class = HistorialPagination 

    def get_queryset(self):
        qs = HistorialMovimientos.objects.all()
        return filtrar_por_sucursal_y_empresa(qs, self.request).order_by('-fecha_hora_movimiento')

class DashboardStatsAPI(APIView):
    permission_classes = [permissions.IsAuthenticated, IsJefeDeBodega] # Asumo que IsJefeDeBodega ya lo tienes importado

    def get(self, request, format=None):
        empresa = get_empresa_from_user(request)
        
        qs_mercancia = filtrar_por_sucursal_y_empresa(Mercancia.activos.all(), request)
        qs_despachos = filtrar_por_sucursal_y_empresa(Despacho.objects.filter(activo=True), request)
        qs_ubicaciones = filtrar_por_sucursal_y_empresa(Ubicacion.activos.all(), request)
        
        total_en_bodega = qs_mercancia.filter(estado='En Bodega').count()

        despachos_activos = qs_despachos.filter(
            estado_despacho__in=['Programado', 'En Carga', 'En Tránsito']
        ).count()

        total_ubicaciones = qs_ubicaciones.count()
        ubicaciones_ocupadas = qs_ubicaciones.filter(estado_ocupado=True).count()
        
        porcentaje_ocupacion = 0
        if total_ubicaciones > 0:
            porcentaje_ocupacion = round((ubicaciones_ocupadas / total_ubicaciones) * 100, 1)

        total_historico = qs_mercancia.count()

        estado_counts = qs_mercancia.values('estado').annotate(cantidad=Count('estado'))
        distribution_data = [
            {'name': item['estado'], 'value': item['cantidad']} 
            for item in estado_counts
        ]

        MESES_ESP = {
            1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr', 5: 'May', 6: 'Jun',
            7: 'Jul', 8: 'Ago', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'
        }
        
        movements_data = []
        hoy = timezone.now().date()
        
        for i in range(5, -1, -1):
            mes_target = hoy.month - i
            ano_target = hoy.year
            
            if mes_target <= 0:
                mes_target += 12
                ano_target -= 1
            total_mes = qs_despachos.filter(
                fecha_programada__year=ano_target,
                fecha_programada__month=mes_target
            ).count()
            
            movements_data.append({
                'mes': MESES_ESP[mes_target],
                'despachos': total_mes
            })

        proximos_despachos = qs_despachos.filter(
            estado_despacho__in=['Programado', 'En Carga']
        ).order_by('fecha_programada')[:5]
        
        despachos_list = []
        for d in proximos_despachos:
            bultos = qs_mercancia.filter(id_despacho=d).count()
            despachos_list.append({
                'id': d.id_despacho,
                'ruta': getattr(d.id_ruta, 'nombre_ruta', 'Sin Ruta'), # Protegemos por si es null
                'camion': getattr(d.id_camion, 'patente', 'Sin Asignar'),
                'fecha': d.fecha_programada.strftime('%d/%m') if d.fecha_programada else 'N/A',
                'estado': d.estado_despacho,
                'bultos': bultos
            })

        top_clientes_qs = qs_mercancia.values('id_cliente__nombre_cliente').annotate(cantidad=Count('id_mercancia')).order_by('-cantidad')[:5]
        top_clientes = [
            {'name': item['id_cliente__nombre_cliente'] or 'Sin Cliente', 'value': item['cantidad']}
            for item in top_clientes_qs
        ]

        top_destinos_qs = qs_mercancia.exclude(id_destino__isnull=True).values('id_destino__nombre_ciudad').annotate(cantidad=Count('id_mercancia')).order_by('-cantidad')[:5]
        top_destinos = [
            {'name': item['id_destino__nombre_ciudad'], 'value': item['cantidad']}
            for item in top_destinos_qs
        ]

        totales = qs_mercancia.filter(estado='En Bodega').aggregate(
            total_kg=Sum('kg'),
            total_m3=Sum('m3'),
            valor_total=Sum('precio_total') 
        )
        total_kg = totales['total_kg'] or 0
        total_m3 = totales['total_m3'] or 0
        valor_total = totales['valor_total'] or 0

        alertas = []
        
        if porcentaje_ocupacion >= 90:
            alertas.append({"tipo": "critico", "titulo": "Bodega Saturada", "mensaje": f"La capacidad ha alcanzado un {porcentaje_ocupacion}%. Se requiere liberar espacio urgente."})
        elif porcentaje_ocupacion >= 75:
            alertas.append({"tipo": "advertencia", "titulo": "Alta Ocupación", "mensaje": f"La ocupación está en {porcentaje_ocupacion}%. Planee despachos pronto."})
            
        despachos_atrasados = qs_despachos.filter(estado_despacho='Programado', fecha_programada__lt=timezone.now().date()).count()
        if despachos_atrasados > 0:
            alertas.append({"tipo": "critico", "titulo": "Despachos Atrasados", "mensaje": f"Existen {despachos_atrasados} despachos programados para fechas pasadas que no han salido."})

        data = {
            'metrics': {
                'en_bodega': total_en_bodega,
                'despachos_activos': despachos_activos,
                'ocupacion': porcentaje_ocupacion,
                'total_historico': total_historico,
                'total_kg': total_kg,
                'total_m3': total_m3,
                'valor_total': valor_total 
            },
            'alertas': alertas,
            'distribution_data': distribution_data,
            'movements_data': movements_data,
            'despachos_list': despachos_list,
            'top_clientes': top_clientes,
            'top_destinos': top_destinos,
        }
        
        return Response(data)

class AreaRestringidaListCreateAPI(generics.ListCreateAPIView):
    serializer_class = AreaRestringidaSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return filtrar_por_sucursal_y_empresa(AreaRestringida.objects.all(), self.request)

    def perform_create(self, serializer):
        empresa = get_empresa_from_user(self.request)
        user = self.request.user
        sucursal_empleado = user.perfil.sucursal if hasattr(user, 'perfil') else None 
        serializer.save(empresa=empresa, sucursal=sucursal_empleado)

class AreaRestringidaDetailAPI(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = AreaRestringidaSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return filtrar_por_sucursal_y_empresa(AreaRestringida.objects.all(), self.request)

#Proveedores
class ProveedorListCreateAPI(generics.ListCreateAPIView):
    queryset = Proveedor.activos.all()
    serializer_class = ProveedorSerializer

    def get_queryset(self):
        empresa = get_empresa_from_user(self.request)
        return Proveedor.objects.filter(empresa=empresa)

    def perform_create(self, serializer):
        empresa = get_empresa_from_user(self.request)
        instance=serializer.save(activo=True, empresa=empresa)
        registrar_auditoria(
            empresa=empresa,
            usuario=self.request.user,
            sucursal=self.request.user.perfil.sucursal,
            modelo="Proveedores",
            accion="Creación",
            descripcion=f"Se creó el proveedor: {instance.nombre_proveedor}"
        )

class ProveedorRetrieveUpdateDestroyAPI(generics.RetrieveUpdateDestroyAPIView):
    queryset = Proveedor.objects.all()
    serializer_class = ProveedorSerializer
    
    def get_queryset(self):
        empresa = get_empresa_from_user(self.request)
        return Proveedor.objects.filter(empresa=empresa)
    
    def perform_update(self, serializer):
        instance = serializer.save()
        registrar_auditoria(
            empresa=instance.empresa,
            usuario=self.request.user,
            sucursal=self.request.user.perfil.sucursal,
            modelo="Proveedores",
            accion="Edición",
            descripcion=f"Se actualizaron datos del proveedor: {instance.nombre_proveedor}"
        )
    
    def perform_destroy(self, instance):
        instance.activo = False
        instance.save()
        registrar_auditoria(
            empresa=instance.empresa,
            usuario=self.request.user,
            sucursal=self.request.user.perfil.sucursal,
            modelo="Proveedores",
            accion="Eliminación",
            descripcion=f"Se eliminó (lógico) al proovedor: {instance.nombre_proveedor}"
        )

class RamplaListCreateAPI(generics.ListCreateAPIView):
    queryset = Rampla.activos.all()
    serializer_class = RamplaSerializer

class RamplaRetrieverUpdateDestroyAPI(generics.RetrieveUpdateDestroyAPIView):
    queryset = Rampla.activos.all()
    serializer_class = RamplaSerializer

class RamplaListCreateAPI(generics.ListCreateAPIView):
    serializer_class = RamplaSerializer 
    permission_classes = [permissions.IsAuthenticated, IsJefeDeBodega]
    
    def get_queryset(self):
        empresa = get_empresa_from_user(self.request)
        return Rampla.objects.filter(empresa=empresa, activo=True)
    
    def perform_create(self, serializer):
        empresa = get_empresa_from_user(self.request)
        instance = serializer.save(empresa=empresa, activo=True)
        registrar_auditoria(
            empresa=empresa,
            usuario=self.request.user,
            sucursal=self.request.user.perfil.sucursal,
            modelo="Rampla",
            accion="Creación",
            descripcion=f"Se creó la rampla: {instance.patente}"
        )

class RamplaDetailAPI(generics.RetrieveUpdateDestroyAPIView):
    queryset = Rampla.objects.all() 
    serializer_class = RamplaSerializer
    permission_classes = [permissions.IsAuthenticated, IsJefeDeBodega]

    def get_queryset(self):
        empresa = get_empresa_from_user(self.request)
        return Rampla.objects.filter(empresa=empresa)
    
    def perform_update(self, serializer):
        instance = serializer.save()
        registrar_auditoria(
            empresa=instance.empresa,
            usuario=self.request.user,
            sucursal=self.request.user.perfil.sucursal,
            modelo="Rampla",
            accion="Edición",
            descripcion=f"Se actualizaron datos de la rampla: {instance.patente}"
        )
    
    def perform_destroy(self, instance):
        instance.activo = False
        instance.save()
        registrar_auditoria(
            empresa=instance.empresa,
            usuario=self.request.user,
            sucursal=self.request.user.perfil.sucursal,
            modelo="Rampla",
            accion="Eliminación",
            descripcion=f"Se eliminó la rampla: {instance.patente}"
        )



def formatear_rut(rut_raw):
    if not rut_raw:
        return ""
    rut_limpio = ''.join(c for c in str(rut_raw) if c.isalnum()).upper()
    if len(rut_limpio) < 2:
        return rut_limpio
    cuerpo = rut_limpio[:-1]
    dv = rut_limpio[-1]
    try:
        cuerpo_fmt = f"{int(cuerpo):,}".replace(',', '.')
        return f"{cuerpo_fmt}-{dv}"
    except ValueError:
        return rut_limpio

class GenerarHojaRutaExcelAPI(generics.RetrieveAPIView):
    permission_classes = [permissions.IsAuthenticated]
    lookup_field = 'id_despacho' 

    def get_queryset(self):
        return filtrar_por_sucursal_y_empresa(Despacho.objects.filter(activo=True), self.request)

    def retrieve(self, request, *args, **kwargs):
        despacho = self.get_object()
        empresa = despacho.empresa

        mercancias = Mercancia.activos.filter(id_despacho=despacho).order_by(
            'id_destino__nombre_ciudad', 
            'id_cliente__nombre_cliente'  
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Ruta_{despacho.id_ruta}"

        # --- ESTILOS ---
        bold_font = Font(bold=True)
        center_aligned = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        left_aligned = Alignment(horizontal="left", vertical="center")
        right_aligned = Alignment(horizontal="right", vertical="center")
        fill_ciudad = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        fill_totales = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")

        # --- CABECERA (Ajustada para 8 columnas: de la A a la H) ---
        ws.merge_cells('A1:B4')
        if empresa.logo:
            try:
                img = ExcelImage(empresa.logo.path)
                img.width = 140
                img.height = 70
                ws.add_image(img, 'A1')
            except Exception:
                ws['A1'] = empresa.nombre_empresa
                ws['A1'].font = Font(bold=True, size=16)
                ws['A1'].alignment = center_aligned
        else:
            ws['A1'] = empresa.nombre_empresa
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = center_aligned
        

        ws.merge_cells('C1:F4')
        texto_con_saltos = (
            "HOJA DE RUTA\n"
            "SOCIEDAD COMERCIAL Y TRANSPORTES\n"
            "MEDALLA Y VARGAS LIMITADA\n"
            "R.U.T 76.203.747-5"
            )
        ws['C1'] = texto_con_saltos
        ws['C1'].font = Font(bold=True, size=8)
        ws['C1'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

        ruta_obj = despacho.id_ruta
        if ruta_obj:
            codigo = getattr(ruta_obj, 'codigo_ruta', '') or ''
            nombre = getattr(ruta_obj, 'nombre_ruta', '') or ''
            ws['H3'] = f"RUTA: {codigo}".strip(" -")
            ws['H3'].font = Font(bold=True, size=12)
            ws['H3'].alignment = center_aligned
        else:
            ws['H3'] = "N/A"

        conductor = despacho.id_conductor
        camion = despacho.id_camion
        rampla = despacho.id_rampla

        ws['A6'] = "Fecha:"
        ws['B6'] = despacho.fecha_salida_real.strftime('%d/%m/%Y') if despacho.fecha_programada else ""
        ws['D6'] = "Pte Camión:"
        ws['E6'] = getattr(camion, 'patente', '')

        ws['A7'] = "Conductor:"
        ws['B7'] = f"{getattr(conductor, 'nombre_completo', '')} {getattr(conductor, 'apellido', '')}" if conductor else ''
        ws['D7'] = "Pte Rampla:"
        ws['E7'] = getattr(rampla, 'patente', 'N/A')

        ws['A8'] = "Rut:"
        ws['B8'] = formatear_rut(getattr(conductor, 'rut_conductor', '') if conductor else '') 
        ws['D8'] = "Celular:"
        ws['E8'] = getattr(conductor, 'telefono', '') if conductor else ''

        # --- TABLA DE MERCANCÍAS ---
        headers = ["Cliente", "Factura", "Cant/Tipo", "Kilos", "CodigoI", "Proveedor", "N° Orden", "Valor"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=10, column=col_num)
            cell.value = header
            cell.font = bold_font
            cell.alignment = center_aligned
            cell.border = thin_border

        row_num = 11
        ciudad_actual = None 
        
        suma_kilos = 0.0
        suma_valor = 0.0
        colores_destinos = ["E6F2FF", "E6FFE6", "FFFFE6", "FFE6E6", "F2E6FF", "E6FFFF"]
        indice_color = -1  

        for m in mercancias:
            ciudad_item = getattr(m.id_destino, 'nombre_ciudad', 'Sin Ciudad')

            if ciudad_item != ciudad_actual:
                indice_color = (indice_color + 1) % len(colores_destinos)
                color_actual = colores_destinos[indice_color]
                relleno_encabezado = PatternFill(start_color=color_actual, end_color=color_actual, fill_type="solid")

                ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
                cell = ws.cell(row=row_num, column=1)
                cell.value = f"DESTINO: {ciudad_item.upper()}"
                cell.font = Font(bold=True, size=12, color="000000")
                cell.alignment = left_aligned
                
                for col in range(1, 9):
                    celda_encabezado = ws.cell(row=row_num, column=col)
                    celda_encabezado.border = thin_border
                    celda_encabezado.fill = relleno_encabezado
                    
                row_num += 1 
                ciudad_actual = ciudad_item

            kilos_fila = float(m.kg) if m.kg else 0.0
            valor_fila = float(m.precio_total) if m.precio_total else 0.0
            
            suma_kilos += kilos_fila
            suma_valor += valor_fila
            
            cant = m.cantidad_bultos or 1
            tipo = m.tipo or ''
            cantidad_y_tipo = f"{cant} {tipo}".strip()

            datos_fila = [
                getattr(m.id_cliente, 'nombre_cliente', ''),
                m.factura,
                cantidad_y_tipo, 
                kilos_fila, 
                m.codigo_interno or '',
                getattr(m.id_proveedor, 'nombre_proveedor', '') if m.id_proveedor else '',
                m.numero_orden_entrega or 'Sin N/O', 
                valor_fila
            ]

            for col_num, dato in enumerate(datos_fila, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = dato
                cell.border = thin_border
                
                if col_num in [1, 6]:
                    cell.alignment = left_aligned
                elif col_num in [4, 8]:
                    cell.alignment = right_aligned
                else:
                    cell.alignment = center_aligned
                
                if col_num == 8 and isinstance(dato, (int, float)):
                    cell.number_format = '"$"#,##0' 
            
            row_num += 1

        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
        lbl_totales = ws.cell(row=row_num, column=1)
        lbl_totales.value = "TOTALES GLOBALES DE LA RUTA:"
        lbl_totales.font = Font(bold=True)
        lbl_totales.alignment = Alignment(horizontal="right", vertical="center")
        lbl_totales.fill = fill_totales
        
        for col in range(1, 4):
            ws.cell(row=row_num, column=col).border = thin_border

        cell_kg = ws.cell(row=row_num, column=4)
        cell_kg.value = suma_kilos
        cell_kg.font = Font(bold=True)
        cell_kg.border = thin_border
        cell_kg.fill = fill_totales
        cell_kg.number_format = '#,##0 "kg"' 

        for col in range(5, 8):
            c_vacia = ws.cell(row=row_num, column=col)
            c_vacia.border = thin_border
            c_vacia.fill = fill_totales

        cell_val = ws.cell(row=row_num, column=8)
        cell_val.value = suma_valor
        cell_val.font = Font(bold=True)
        cell_val.border = thin_border
        cell_val.fill = fill_totales
        cell_val.number_format = '"$"#,##0' 

        # --- AJUSTE DE ANCHO DE COLUMNAS ---
        column_widths = {'A': 15, 'B': 15, 'C': 18, 'D': 12, 'E': 15, 'F': 20, 'G': 15, 'H': 15}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = False
        
        ws.page_setup.paperSize = 1 
        
        ws.print_options.horizontalCentered = True
        
        ws.page_margins = PageMargins(
            left=0.35,   
            right=0.35, 
            top=0.75,   
            bottom=0.75, 
            header=0.3,
            footer=0.3
        )

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Ruta_{despacho.id_ruta}.xlsx"'
        wb.save(response)
        
        return response
    

class GenerarOrdenesDespachoAPI(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, id_despacho):
        empresa = get_empresa_from_user(request)
        despacho = get_object_or_404(Despacho, id_despacho=id_despacho, empresa=empresa)
        try:
            generar_numeros_orden_despacho(despacho)
            return Response({"mensaje": "Órdenes generadas exitosamente."}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        


class CotizacionListCreateAPI(generics.ListCreateAPIView):
    serializer_class = CotizacionSerializer

    def get_queryset(self):
        empresa = get_empresa_from_user(self.request)
        return Cotizacion.objects.filter(empresa=empresa, activo=True).order_by('-fecha_creacion')

    def perform_create(self, serializer):
        empresa = get_empresa_from_user(self.request)
        
        instance = serializer.save(
            empresa=empresa,
            id_usuario_creacion=self.request.user,
            estado_cotizacion='En proceso'
        )
        
        registrar_auditoria(
            empresa=empresa,
            usuario=self.request.user,
            sucursal=self.request.user.perfil.sucursal,
            modelo="Cotizaciones",
            accion="Creación",
            descripcion=f"Se creó cotización para: {instance.nombre_cliente}"
        )


class CotizacionRetrieveUpdateDestroyAPI(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = CotizacionSerializer

    def get_queryset(self):
        empresa = get_empresa_from_user(self.request)
        return Cotizacion.objects.filter(empresa=empresa)

    def perform_update(self, serializer):
        instance = self.get_object()
        nuevo_estado = serializer.validated_data.get('estado_cotizacion', instance.estado_cotizacion)

        if nuevo_estado == 'Cotizado' and instance.estado_cotizacion != 'Cotizado':
            serializer.save(fecha_confirmacion=timezone.now())
        else:
            serializer.save()

        instance = serializer.instance 
        
        registrar_auditoria(
            empresa=instance.empresa,
            usuario=self.request.user,
            sucursal=self.request.user.perfil.sucursal,
            modelo="Cotizaciones",
            accion="Edición",
            descripcion=f"Cotización de {instance.nombre_cliente} actualizada a estado: {instance.estado_cotizacion}"
        )

    def perform_destroy(self, instance):
        instance.activo = False
        instance.save()
        
        registrar_auditoria(
            empresa=instance.empresa,
            usuario=self.request.user,
            sucursal=self.request.user.perfil.sucursal,
            modelo="Cotizaciones",
            accion="Eliminación",
            descripcion=f"Se eliminó (lógico) la cotización de: {instance.nombre_cliente}"
        )


class InvitarColaboradorAPI(generics.CreateAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = InvitacionColaboradorSerializer 

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        usuario_que_invita = request.user
        despacho_id = self.kwargs.get('id_despacho') 
        despacho = get_object_or_404(Despacho, pk=despacho_id)
        usuario_invitado_id = serializer.validated_data['usuario_invitado_id']
        usuario_invitado = get_object_or_404(User, pk=usuario_invitado_id)

        if usuario_invitado.perfil.sucursal == despacho.sucursal:
            return Response(
                {"error": "El usuario ya pertenece a la sucursal de este despacho y tiene acceso por defecto."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if usuario_que_invita.perfil.rol != 'DUENO' and despacho.sucursal != usuario_que_invita.perfil.sucursal:
            return Response(
                {"error": "No tienes permiso para invitar colaboradores a un despacho que no pertenece a tu sucursal."}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        empresa_actual = get_empresa_from_user(request)
        if despacho.empresa != empresa_actual:
            return Response(
                {"error": "No puedes compartir un despacho de otra empresa."}, 
                status=status.HTTP_403_FORBIDDEN
            )

        invitacion, created = PermisoColaboracion.objects.get_or_create(
            despacho=despacho,
            usuario_invitado=usuario_invitado,
            defaults={
                'otorgado_por': usuario_que_invita,
                'activo': True
            }
        )

        if not created:
            if not invitacion.activo:
                invitacion.activo = True
                invitacion.otorgado_por = usuario_que_invita
                invitacion.save()
                return Response(
                    {"mensaje": f"Se reactivó el acceso a {usuario_invitado.username}."}, 
                    status=status.HTTP_200_OK
                )
            else:
                return Response(
                    {"mensaje": f"{usuario_invitado.username} ya tiene acceso a este despacho."}, 
                    status=status.HTTP_200_OK
                )

        return Response({
            "mensaje": f"Se otorgó permiso exitosamente a {usuario_invitado.username} para el despacho #{despacho.id_despacho}."
        }, status=status.HTTP_201_CREATED)
#DJANGO METODO SIN REACT (FUNCIONAL)


# @login_required
# def panel_inventario(request):
#     """
#     Vista principal de la app inventario, muestra un resumen.
#     """
#     total_en_bodega = Mercancia.objects.filter(estado='En Bodega').count()
#     despachos_programados = Despacho.objects.filter(estado_despacho='Programado').count()
#     ubicaciones_libres = Ubicacion.objects.filter(estado_ocupado=False).count()
#     total_clientes = Cliente.objects.count()

#     context = {
#         'total_en_bodega': total_en_bodega,
#         'despachos_programados': despachos_programados,
#         'ubicaciones_libres': ubicaciones_libres,
#         'total_clientes': total_clientes,
#     }
#     return render(request, 'inventario/panel.html', context)
#     
# # --- CRUD Completo para Mercancia ---

# class MercanciaListView(LoginRequiredMixin, ListView):
#     model = Mercancia
#     template_name = 'inventario/mercancia_list.html'
#     context_object_name = 'mercancias'
#     queryset = Mercancia.objects.select_related('id_cliente', 'id_ubicacion_actual', 'id_destino').order_by('-fecha_ingreso')

# class MercanciaDetailView(LoginRequiredMixin, DetailView):
#     model = Mercancia
#     template_name = 'inventario/mercancia_detail.html'
#     context_object_name = 'mercancia'

# class MercanciaCreateView(LoginRequiredMixin, CreateView):
#     model = Mercancia
#     template_name = 'inventario/mercancia_form.html'
#     fields = [
#         'id_cliente', 'descripcion_carga', 'cantidad_bultos', 
#         'kg', 'm3', 'id_ubicacion_actual', 'id_destino'
#     ]
#     success_url = reverse_lazy('mercancia-list')

#     def form_valid(self, form):
#         form.instance.id_usuario_creacion = self.request.user
#         return super().form_valid(form)

#     def form_invalid(self, form):
#         print("="*20, "FORMULARIO INVÁLIDO", "="*20)
#         print(form.errors.as_json())
#         print("="*50)
#         return super().form_invalid(form)

# class MercanciaUpdateView(LoginRequiredMixin, UpdateView):
#     model = Mercancia
#     template_name = 'inventario/mercancia_form.html'
#     fields = [
#         'id_cliente', 'descripcion_carga', 'cantidad_bultos', 
#         'kg', 'm3', 'id_ubicacion_actual', 'id_destino', 
#         'estado', 'id_despacho'
#     ]
    
#     def form_valid(self, form):
#         form.instance.id_usuario_ultima_modificacion = self.request.user
#         return super().form_valid(form)
    
#     def get_success_url(self):
#         return reverse_lazy('mercancia-detail', kwargs={'pk': self.object.pk})

# class MercanciaDeleteView(LoginRequiredMixin, DeleteView):
#     model = Mercancia
#     template_name = 'inventario/mercancia_confirm_delete.html'
#     success_url = reverse_lazy('mercancia-list')


# # --- CRUD para Clientes ---

# class ClienteListView(LoginRequiredMixin, ListView):
#     model = Cliente
#     template_name = 'inventario/cliente_list.html'
#     context_object_name = 'clientes'

# class ClienteCreateView(LoginRequiredMixin, CreateView):
#     model = Cliente
#     template_name = 'inventario/cliente_form.html'
#     fields = ['nombre_cliente', 'rut_cliente', 'telefono_contacto', 'email_contacto']
#     success_url = reverse_lazy('cliente-list')

# class ClienteUpdateView(LoginRequiredMixin, UpdateView):
#     model = Cliente
#     template_name = 'inventario/cliente_form.html'
#     fields = ['nombre_cliente', 'rut_cliente', 'telefono_contacto', 'email_contacto']
#     success_url = reverse_lazy('cliente-list')

# class ClienteDeleteView(LoginRequiredMixin, DeleteView):
#     model = Cliente
#     template_name = 'inventario/cliente_confirm_delete.html'
#     success_url = reverse_lazy('cliente-list')


# # --- CRUD para Despachos ---

# class DespachoForm(forms.ModelForm):
#     class Meta:
#         model = Despacho
#         fields = [
#             'fecha_programada', 'fecha_salida_real', 'id_camion', 
#             'id_conductor', 'id_ruta', 'estado_despacho'
#         ]
        
#         widgets = {
#             'fecha_programada': forms.DateInput(
#                 attrs={'type': 'date', 'class': 'form-control'}
#             ),
#             'fecha_salida_real': forms.DateTimeInput(
#                 attrs={'type': 'datetime-local', 'class': 'form-control'}
#             ),
#             'id_camion': forms.Select(attrs={'class': 'form-select'}),
#             'id_conductor': forms.Select(attrs={'class': 'form-select'}),
#             'id_ruta': forms.Select(attrs={'class': 'form-select'}),
#             'estado_despacho': forms.Select(attrs={'class': 'form-select'}),
#         }

# class DespachoListView(LoginRequiredMixin, ListView):
#     model = Despacho
#     template_name = 'inventario/despacho_list.html'
#     context_object_name = 'despachos'
#     queryset = Despacho.objects.select_related('id_camion', 'id_conductor', 'id_ruta').order_by('-fecha_programada')

# class DespachoDetailView(LoginRequiredMixin, DetailView):
#     model = Despacho
#     template_name = 'inventario/despacho_detail.html'
#     context_object_name = 'despacho'
    
#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['mercancias_asignadas'] = Mercancia.objects.filter(id_despacho=self.object)
#         return context

# class DespachoCreateView(LoginRequiredMixin, CreateView):
#     model = Despacho
#     form_class = DespachoForm  
#     template_name = 'inventario/despacho_form.html'

#     success_url = reverse_lazy('despacho-list')

#     def get_form(self, form_class=None):
#         form = super().get_form(form_class)
#         form.fields.pop('fecha_salida_real', None)
#         return form

#     def form_valid(self, form):
#         form.instance.id_usuario_creacion = self.request.user
#         return super().form_valid(form)

# class DespachoUpdateView(LoginRequiredMixin, UpdateView):
#     model = Despacho
#     form_class = DespachoForm 
#     template_name = 'inventario/despacho_form.html'

    
#     def form_valid(self, form):
#         form.instance.id_usuario_ultima_modificacion = self.request.user
#         return super().form_valid(form)

#     def get_success_url(self):
#         return reverse_lazy('despacho-detail', kwargs={'pk': self.object.pk})

# class DespachoDeleteView(LoginRequiredMixin, DeleteView):
#     model = Despacho
#     template_name = 'inventario/despacho_confirm_delete.html'
#     success_url = reverse_lazy('despacho-list')


# # --- CRUD para Conductores ---

# class ConductorListView(LoginRequiredMixin, ListView):
#     model = Conductor
#     template_name = 'inventario/conductor_list.html'
#     context_object_name = 'conductores'

# class ConductorCreateView(LoginRequiredMixin, CreateView):
#     model = Conductor
#     template_name = 'inventario/conductor_form.html'
#     fields = ['nombre_completo', 'rut_conductor', 'numero_licencia', 'telefono']
#     success_url = reverse_lazy('conductor-list')

# class ConductorUpdateView(LoginRequiredMixin, UpdateView):
#     model = Conductor
#     template_name = 'inventario/conductor_form.html'
#     fields = ['nombre_completo', 'rut_conductor', 'numero_licencia', 'telefono']
#     success_url = reverse_lazy('conductor-list')

# class ConductorDeleteView(LoginRequiredMixin, DeleteView):
#     model = Conductor
#     template_name = 'inventario/conductor_confirm_delete.html'
#     success_url = reverse_lazy('conductor-list')


# # --- CRUD para Camiones ---

# class CamionListView(LoginRequiredMixin, ListView):
#     model = Camion
#     template_name = 'inventario/camion_list.html'
#     context_object_name = 'camiones'

# class CamionCreateView(LoginRequiredMixin, CreateView):
#     model = Camion
#     template_name = 'inventario/camion_form.html'
#     fields = ['patente', 'marca', 'modelo', 'capacidad_max_kg', 'capacidad_max_m3']
#     success_url = reverse_lazy('camion-list')

# class CamionUpdateView(LoginRequiredMixin, UpdateView):
#     model = Camion
#     template_name = 'inventario/camion_form.html'
#     fields = ['patente', 'marca', 'modelo', 'capacidad_max_kg', 'capacidad_max_m3']
#     success_url = reverse_lazy('camion-list')

# class CamionDeleteView(LoginRequiredMixin, DeleteView):
#     model = Camion
#     template_name = 'inventario/camion_confirm_delete.html'
#     success_url = reverse_lazy('camion-list')


# # --- CRUD para Rutas ---

# class RutaListView(LoginRequiredMixin, ListView):
#     model = Ruta
#     template_name = 'inventario/ruta_list.html'
#     context_object_name = 'rutas'

# class RutaCreateView(LoginRequiredMixin, CreateView):
#     model = Ruta
#     template_name = 'inventario/ruta_form.html'
#     fields = ['nombre_ruta', 'descripcion']
#     success_url = reverse_lazy('ruta-list')

# class RutaUpdateView(LoginRequiredMixin, UpdateView):
#     model = Ruta
#     template_name = 'inventario/ruta_form.html'
#     fields = ['nombre_ruta', 'descripcion']
#     success_url = reverse_lazy('ruta-list')

# class RutaDeleteView(LoginRequiredMixin, DeleteView):
#     model = Ruta
#     template_name = 'inventario/ruta_confirm_delete.html'
#     success_url = reverse_lazy('ruta-list')


# # --- CRUD para Destinos ---

# class DestinoListView(LoginRequiredMixin, ListView):
#     model = Destino
#     template_name = 'inventario/destino_list.html'
#     context_object_name = 'destinos'

# class DestinoCreateView(LoginRequiredMixin, CreateView):
#     model = Destino
#     template_name = 'inventario/destino_form.html'
#     fields = ['nombre_ciudad', 'region']
#     success_url = reverse_lazy('destino-list')

# class DestinoUpdateView(LoginRequiredMixin, UpdateView):
#     model = Destino
#     template_name = 'inventario/destino_form.html'
#     fields = ['nombre_ciudad', 'region']
#     success_url = reverse_lazy('destino-list')

# class DestinoDeleteView(LoginRequiredMixin, DeleteView):
#     model = Destino
#     template_name = 'inventario/destino_confirm_delete.html'
#     success_url = reverse_lazy('destino-list')


# # --- CRUD para Ubicaciones ---

# class UbicacionListView(LoginRequiredMixin, ListView):
#     model = Ubicacion
#     template_name = 'inventario/ubicacion_list.html'
#     context_object_name = 'ubicaciones'

# class UbicacionCreateView(LoginRequiredMixin, CreateView):
#     model = Ubicacion
#     template_name = 'inventario/ubicacion_form.html'
#     fields = [
#         'codigo_ubicacion', 'pasillo', 'estanteria', 'nivel', 
#         'pos_x', 'pos_y', 'pos_z', 
#         'capacidad_max_kg', 'capacidad_max_m3', 'estado_ocupado'
#     ]
#     success_url = reverse_lazy('ubicacion-list')

# class UbicacionUpdateView(LoginRequiredMixin, UpdateView):
#     model = Ubicacion
#     template_name = 'inventario/ubicacion_form.html'
#     fields = [
#         'codigo_ubicacion', 'pasillo', 'estanteria', 'nivel', 
#         'pos_x', 'pos_y', 'pos_z', 
#         'capacidad_max_kg', 'capacidad_max_m3', 'estado_ocupado'
#     ]
#     success_url = reverse_lazy('ubicacion-list')

# class UbicacionDeleteView(LoginRequiredMixin, DeleteView):
#     model = Ubicacion
#     template_name = 'inventario/ubicacion_confirm_delete.html'
#     success_url = reverse_lazy('ubicacion-list')