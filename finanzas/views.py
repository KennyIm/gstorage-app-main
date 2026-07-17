from django.shortcuts import render, get_object_or_404
from rest_framework import generics, status, permissions
from cryptography.fernet import Fernet
from django.conf import settings
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from django.utils.dateparse import parse_date
from django.db import transaction
from decimal import Decimal
from datetime import timedelta, datetime
import datetime
from django.utils import timezone
from django.db.models import Max, Sum, Q, CharField
from django.db.models.functions import ExtractMonth, ExtractYear
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from cryptography.fernet import Fernet
from django.conf import settings

from .pagination import FinanzasPendientesPagination, GastosOperativosPagination
from .models import DocumentoCobro, PagoRecibido, GastoOperativo, ProveedorGasto
from .serializers import MercanciaPendienteCobroSerializer, DocumentoCobroSerializer, DocumentoCobroListSerializer, RegistrarPagoSerializer, GastoOperativoSerializer, ProveedorGastoSerializer,DocumentoCobroDashboardSerializer
from inventario.models import Mercancia, Cliente, Despacho
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from inventario.views import filtrar_por_sucursal_y_empresa, formatear_rut
from django.db.models.functions import Cast

class MercanciasPendientesListaAPI(generics.ListAPIView):
    serializer_class = MercanciaPendienteCobroSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = None
    def get_queryset(self):
        empresa_id = self.request.user.perfil.empresa_id
        cliente_id = self.request.query_params.get('cliente_id')        
        if not cliente_id:
            return Mercancia.objects.none()
        return Mercancia.objects.filter(
            empresa_id=empresa_id,
            id_cliente_id=cliente_id,
            estado_cobranza='Pendiente'
        ).order_by('fecha_ingreso')


class GenerarCobroGenericView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    queryset = DocumentoCobro.objects.all()
    serializer_class = DocumentoCobroSerializer

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        datos = request.data
        
        if hasattr(datos, 'getlist'):
            raw_ids = datos.getlist('mercancias_ids') or datos.getlist('mercancias_ids[]')
        else:
            raw_ids = datos.get('mercancias_ids') or datos.get('mercancias_ids[]')

        if not raw_ids:
            return Response({"error": "Faltan las mercancías a facturar."}, status=status.HTTP_400_BAD_REQUEST)
            
        if not isinstance(raw_ids, list):
            if isinstance(raw_ids, str) and ',' in raw_ids:
                raw_ids = raw_ids.split(',')
            else:
                raw_ids = [raw_ids]
                
        mercancias_ids = list(set([int(i) for i in raw_ids if i and str(i).strip()]))
        cliente_id = datos.get('cliente_id')
        tipo_documento = datos.get('tipo_documento', 'Factura')
        condicion_pago = datos.get('condicion_pago', 'Dias_30')
        numero_documento = datos.get('numero_documento')
        fecha_emision_str = datos.get('fecha_emision')
        
        if fecha_emision_str:
            naive_datetime = datetime.datetime.strptime(fecha_emision_str, '%Y-%m-%d')
            fecha_emision = timezone.make_aware(naive_datetime)
        else:
            fecha_emision = timezone.now()
            
        archivo_pdf = request.FILES.get('pdf_documento')

        if not mercancias_ids or not cliente_id:
            return Response({"error": "Faltan datos obligatorios: cliente o mercancías."}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            empresa = request.user.perfil.empresa
            sucursal = request.user.perfil.sucursal
            cliente = Cliente.objects.get(id_cliente=cliente_id, empresa=empresa)
            
            if tipo_documento == 'Guia_Cobro' and not numero_documento:
                ultima_guia = self.get_queryset().filter(
                    empresa=empresa,
                    tipo_documento='Guia_Cobro'
                ).order_by('-numero_documento').only('numero_documento').first()

                numero_documento = (ultima_guia.numero_documento + 1) if (ultima_guia and ultima_guia.numero_documento) else 1
            
            mercancias_lista = list(
                Mercancia.objects.select_for_update()
                .filter(
                    id_mercancia__in=mercancias_ids, 
                    id_cliente=cliente,
                    estado_cobranza='Pendiente' 
                )
                .only('id_mercancia', 'precio_total', 'paga_proveedor', 'id_proveedor_id')
            )

            if len(mercancias_lista) != len(mercancias_ids):
                ids_encontrados = [m.id_mercancia for m in mercancias_lista]
                ids_conflictivos = list(set(mercancias_ids) - set(ids_encontrados))
                
                return Response(
                    {
                        "error": "Conflicto en la selección: Algunas mercancías ya se procesaron (En_Proceso/Facturadas) o corresponden a otro cliente.",
                        "mercancias_conflictivas": ids_conflictivos 
                    }, 
                    status=status.HTTP_400_BAD_REQUEST
                )

            primera_m = mercancias_lista[0]
            el_proveedor_paga = primera_m.paga_proveedor
            proveedor_asignado = primera_m.id_proveedor if el_proveedor_paga else None
            
            subtotal = sum(Decimal(str(m.precio_total)) for m in mercancias_lista)
            iva = Decimal('0.19') * subtotal if tipo_documento == 'Factura' else Decimal(0)
            total = subtotal + iva
            
            dias_plazo = {
                'Contra_Entrega': 0, 'Dias_15': 15, 'Dias_30': 30, 'Dias_45': 45, 'Dias_60': 60
            }.get(condicion_pago, 30)
            vencimiento = fecha_emision + timedelta(days=dias_plazo)
            
            nuevo_documento = DocumentoCobro.objects.create(
                empresa=empresa,
                sucursal=sucursal,
                cliente_deudor=cliente, 
                proveedor_deudor=proveedor_asignado,
                tipo_documento=tipo_documento,
                condicion_pago=condicion_pago,
                fecha_vencimiento=vencimiento,
                fecha_emision=fecha_emision,
                numero_documento=numero_documento,
                pdf_documento=archivo_pdf,
                subtotal=subtotal,
                iva=iva,
                total_a_pagar=total,
                saldo_pendiente=total,
                estado='Emitido'
            )
            
            nuevo_documento.mercancias_asociadas.set(mercancias_ids)
            
            Mercancia.objects.filter(id_mercancia__in=mercancias_ids).update(
                estado_cobranza='En_Proceso', 
                tipo_documento_pago=tipo_documento
            )
            
            documento_listo = self.get_queryset().select_related(
                'cliente_deudor', 'proveedor_deudor'
            ).prefetch_related('mercancias_asociadas').get(pk=nuevo_documento.pk)
            
            serializer = self.get_serializer(documento_listo)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        except Cliente.DoesNotExist:
            return Response({"error": "Cliente no encontrado en los registros de la empresa."}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": f"Error interno en el servidor: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        

class DocumentosEmitidosListaAPIView(generics.ListAPIView):
    serializer_class = DocumentoCobroListSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = FinanzasPendientesPagination

    def get_queryset(self):
        empresa_id = self.request.user.perfil.empresa_id
        queryset = DocumentoCobro.objects.filter(empresa_id=empresa_id, activo=True)
        tab = self.request.query_params.get('tab')
        estado = self.request.query_params.get('estado')
        numero_doc = self.request.query_params.get('numero_documento')
        deudor = self.request.query_params.get('deudor')
        emision_desde = self.request.query_params.get('emision_desde')
        emision_hasta = self.request.query_params.get('emision_hasta')
        vence_desde = self.request.query_params.get('vence_desde')
        vence_hasta = self.request.query_params.get('vence_hasta')

        if tab == 'activos':
            queryset = queryset.exclude(estado='Pagado')
        elif tab == 'pagados':
            queryset = queryset.filter(estado='Pagado')

        if estado and estado != 'Todos':
            queryset = queryset.filter(estado=estado)
        if numero_doc:
            queryset = queryset.annotate(
                num_str=Cast('numero_documento', CharField())
            ).filter(num_str__icontains=numero_doc)
        if deudor:
            queryset = queryset.filter(
                Q(cliente_deudor__nombre_cliente__icontains=deudor) |
                Q(proveedor_deudor__nombre_proveedor__icontains=deudor)
            )
        if emision_desde:
            queryset = queryset.filter(fecha_emision__gte=emision_desde)
        if emision_hasta:
            queryset = queryset.filter(fecha_emision__lte=emision_hasta)
        if vence_desde:
            queryset = queryset.filter(fecha_vencimiento__gte=vence_desde)
        if vence_hasta:
            queryset = queryset.filter(fecha_vencimiento__lte=vence_hasta)

        return queryset.order_by('-fecha_emision', '-id')
    

class RegistrarPagoAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    parser_classes = [MultiPartParser, FormParser]

    @transaction.atomic
    def post(self, request):
        serializer = RegistrarPagoSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        datos = serializer.validated_data
        
        documento_id = datos.get('documento_id')
        monto_pagado = datos.get('monto_pagado')
        medio_pago = datos.get('medio_pago')
        numero_operacion_banco = datos.get('numero_operacion_banco', '').strip()

        if medio_pago == 'Efectivo' and not numero_operacion_banco:
            total_efectivo = PagoRecibido.objects.filter(
                empresa=request.user.perfil.empresa,
                medio_pago='Efectivo'
            ).count()
            numero_operacion_banco = f"CAJA-{total_efectivo + 1:04d}"
        elif medio_pago in ['Cheque_Dia', 'Cheque_Fecha'] and not numero_operacion_banco:
            numero_operacion_banco = "S/N Cheque"
        
        comprobante_archivo = datos.get('comprobante_banco')

        try:
            documento = DocumentoCobro.objects.select_for_update().get(
                id=documento_id,
                empresa=request.user.perfil.empresa
            )

            if documento.saldo_pendiente <= 0:
                return Response({"error": "Este documento ya está pagado en su totalidad."}, status=status.HTTP_400_BAD_REQUEST)
            
            monto_pago = Decimal(str(monto_pagado))

            if monto_pago > documento.saldo_pendiente:
                return Response({"error": f"El pago (${monto_pago}) supera el saldo pendiente (${documento.saldo_pendiente})."}, status=status.HTTP_400_BAD_REQUEST)
            
            PagoRecibido.objects.create(
                empresa=request.user.perfil.empresa,
                documento_pagado=documento,
                monto_pagado=monto_pago,
                medio_pago=medio_pago,
                numero_operacion_banco=numero_operacion_banco,
                comprobante_banco=comprobante_archivo
            )

            documento.saldo_pendiente -= monto_pago

            if documento.saldo_pendiente == 0:
                documento.estado = 'Pagado'
                documento.mercancias_asociadas.update(estado_cobranza='Pagado')
            else:
                documento.estado = 'Abonado'
            
            documento.save()

            return Response({"mensaje": "Pago registrado correctamente", "nuevo_saldo":documento.saldo_pendiente}, status=status.HTTP_200_OK)
        
        except DocumentoCobro.DoesNotExist:
            return Response({"error": "Documento no encontrado."}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class PerfilFinancieroClienteAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request, cliente_id):
        empresa = request.user.perfil.empresa

        try:
            cliente = Cliente.objects.get(id_cliente=cliente_id, empresa=empresa)

            mercancias_sueltas = Mercancia.objects.filter(
                id_cliente=cliente,
                estado_cobranza='Pendiente',
                activo=True
            ).select_related('id_proveedor')

            documentos = DocumentoCobro.objects.filter(
                cliente_deudor=cliente,
                activo=True
            ).select_related('proveedor_deudor').prefetch_related('pagos')

            historial_unificado = []

            monto_sin_facturar = 0
            monto_facturado = 0
            monto_pagado = 0

            for m in mercancias_sueltas:
                monto_sin_facturar += m.precio_total
                historial_unificado.append({
                    'id_unico': f"mer_{m.id_mercancia}",
                    'tipo_registro': 'Sin_Facturar',
                    'numero': m.numero_orden_entrega or f"Int-{m.id_mercancia}",
                    'proveedor_nombre': m.id_proveedor.nombre_proveedor if m.id_proveedor else "N/A",
                    'paga_proveedor': m.paga_proveedor,
                    'fecha': m.fecha_ingreso.strftime('%Y-%m-%d'),
                    'monto': m.precio_total,
                    'estado': 'Sin Facturar'
                })

            for d in documentos:
                fecha_final = d.fecha_emision.strftime('%Y-%m-%d')
                if d.estado == 'Pagado':
                    monto_pagado += d.total_a_pagar
                    ultimo_pago = d.pagos.aggregate(ultimo=Max('fecha_pago'))['ultimo']
                    if ultimo_pago:
                        fecha_final = ultimo_pago.strftime('%Y-%m-%d')
                else:
                    monto_facturado += d.total_a_pagar
                
                historial_unificado.append({
                    'id_unico' : f"doc_{d.id}",
                    'tipo_registro' : 'Documento',
                    'numero' : d.numero_documento or f"Borrador-{d.id}",
                    'proveedor_nombre' : d.proveedor_deudor.nombre_proveedor if d.proveedor_deudor else "N/A",
                    'paga_proveedor' : d.proveedor_deudor is not None,
                    'fecha' : fecha_final,
                    'monto' : d.total_a_pagar,
                    'estado' : d.get_estado_display()
                })
            
            rut_desencriptado = ""
            if cliente.rut_cliente_cifrado:
                try:
                    fernet = Fernet(settings.FIELD_ENCRYPTION_KEY.encode())
                    rut_desencriptado = fernet.decrypt(cliente.rut_cliente_cifrado.encode('utf-8')).decode('utf-8')
                except Exception:
                    rut_desencriptado = "Error al desencriptar"


            payload = {
                'cliente':{
                    'nombre': cliente.nombre_cliente,
                    'rut': rut_desencriptado
                },
                'metricas': {
                    'sin_facturar': float(monto_sin_facturar),
                    'facturado': float(monto_facturado),
                    'pagado': float(monto_pagado),
                },
                'historial': historial_unificado
            }
            return Response(payload, status=status.HTTP_200_OK)
        
        except Cliente.DoesNotExist:
            return Response({"error": "Cliente no encontrado."}, status=status.HTTP_404_NOT_FOUND)

class GastoOperativoListCreateAPIView(generics.ListCreateAPIView):
    serializer_class = GastoOperativoSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    pagination_class = GastosOperativosPagination

    def get_queryset(self):
        empresa_id = self.request.user.perfil.empresa_id
        queryset = GastoOperativo.objects.filter(empresa_id=empresa_id, activo=True)
        tipo = self.request.query_params.get('tipo')
        estado = self.request.query_params.get('estado')
        busqueda = self.request.query_params.get('busqueda')

        if tipo and tipo != 'Todos':
            queryset = queryset.filter(tipo_gasto=tipo)
        if estado and estado != 'Todos':
            queryset = queryset.filter(estado=estado)
        if busqueda:
            queryset = queryset.annotate(
                doc_str=Cast('numero_documento', CharField())
            ).filter(
                Q(descripcion__icontains=busqueda) |
                Q(doc_str__icontains=busqueda) |
                Q(proveedor__nombre_proveedor__icontains=busqueda)
            )

        return queryset.order_by('-fecha_gasto', '-id')
    
class ProveedorGastoListCreateAPIView(generics.ListCreateAPIView): 
    serializer_class = ProveedorGastoSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return ProveedorGasto.objects.filter(
            empresa_id=self.request.user.perfil.empresa_id, 
            activo=True
        ).order_by('nombre_proveedor')

    def perform_create(self, serializer):
        serializer.save(empresa=self.request.user.perfil.empresa)


class PagarGastoOperativoAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    parser_classes = [MultiPartParser, FormParser]
    def patch(self, request, pk):
        try:
            gasto = GastoOperativo.objects.get(
                id=pk, 
                empresa=request.user.perfil.empresa, 
                activo=True
            )
            if gasto.estado == 'Pagado':
                return Response({"error": "Este gasto ya se encuentra totalmente liquidado."}, status=status.HTTP_400_BAD_REQUEST)            
            archivo_respaldo = request.FILES.get('comprobante_adjunto')
            if archivo_respaldo:
                gasto.comprobante_adjunto = archivo_respaldo
            
            gasto.estado = 'Pagado'
            gasto.save()
            
            return Response({"mensaje": "Gasto liquidado con éxito."}, status=status.HTTP_200_OK)
            
        except GastoOperativo.DoesNotExist:
            return Response({"error": "Gasto operativo no encontrado."}, status=status.HTTP_404_NOT_FOUND)
        

class DashboardFinanzasConsolidadoAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        empresa_id = request.user.perfil.empresa_id
        hoy = datetime.date.today()
        documentos_qs = DocumentoCobro.objects.filter(
            empresa_id=empresa_id, 
            activo=True
        ).select_related('cliente_deudor', 'proveedor_deudor').order_by('-fecha_emision')
        
        documentos_serializer = DocumentoCobroDashboardSerializer(documentos_qs, many=True)
        gastos_mensuales = GastoOperativo.objects.filter(
            empresa_id=empresa_id,
            activo=True,
            fecha_gasto__year=hoy.year
        ).annotate(
            mes=ExtractMonth('fecha_gasto')
        ).values('mes').annotate(
            total=Sum('monto_total')
        ).order_by('mes')
        ventas_mensuales = DocumentoCobro.objects.filter(
            empresa_id=empresa_id,
            activo=True,
            fecha_emision__year=hoy.year
        ).annotate(
            mes=ExtractMonth('fecha_emision')
        ).values('mes').annotate(
            facturado=Sum('subtotal', filter=Q(tipo_documento='Factura')),
            no_facturado=Sum('subtotal', filter=Q(tipo_documento='Guia_Cobro'))
        ).order_by('mes')

        tendencias_map = {}
        for m in range(1, 13):
            mes_str = f"{hoy.year}-{str(m).zfill(2)}"
            tendencias_map[m] = {
                "mes": mes_str, 
                "ventas_facturadas": 0,    
                "ventas_por_facturar": 0,
                "compras": 0
            }
        for v in ventas_mensuales:
            m_id = v['mes']
            if m_id in tendencias_map:
                tendencias_map[m_id]['ventas_facturadas'] = float(v['facturado'] or 0)
                tendencias_map[m_id]['ventas_por_facturar'] = float(v['no_facturado'] or 0)
        for g in gastos_mensuales:
            m_id = g['mes']
            if m_id in tendencias_map:
                tendencias_map[m_id]['compras'] = float(g['total'] or 0)

        total_por_pagar = GastoOperativo.objects.filter(
            empresa_id=empresa_id,
            estado='Pendiente',
            activo=True
        ).aggregate(total=Sum('monto_total'))['total'] or 0

        payload = {
            "saldo_bancos": 50000000, 
            "cuentas_por_pagar_exigible": float(total_por_pagar),
            "documentos": documentos_serializer.data,
            "grafico_mensual": list(tendencias_map.values())
        }

        return Response(payload, status=status.HTTP_200_OK)
    


class GenerarFacturasDespachoExcelAPI(generics.RetrieveAPIView):
    permission_classes = [permissions.IsAuthenticated]
    lookup_field = 'id_despacho' 

    def get_queryset(self):
        return filtrar_por_sucursal_y_empresa(Despacho.objects.filter(activo=True), self.request)

    def classify_macro_zone(self, ciudad_nombre):
        """Clasifica las ciudades en las Macro-Zonas operativas de la empresa"""
        ciudad = str(ciudad_nombre).lower()
        if any(x in ciudad for x in ['antof', 'calama', 'mejillon', 'tocopilla']):
            return "ANTOFA"
        if any(x in ciudad for x in ['iquique', 'hospicio']):
            return "IQUIQUE"
        if any(x in ciudad for x in ['copiapo', 'caldera', 'vallenar']):
            return "COPIAPO"
        return "ANTOFA"

    def write_detailed_sheet(self, ws, title_text, mercancias_list, documentos_map, fernet, is_general=True, mercancia_to_general_row=None):
        ws.views.sheetView[0].showGridLines = True

        font_titulo = Font(name="Arial", size=13, bold=True, color="1E3A8A")
        bold_font = Font(name="Arial", size=9, bold=True)
        font_regular = Font(name="Arial", size=9)
        
        center_aligned = Alignment(horizontal="center", vertical="center")
        left_aligned = Alignment(horizontal="left", vertical="center")
        right_aligned = Alignment(horizontal="right", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin', color="DDDDDD"), right=Side(style='thin', color="DDDDDD"),
            top=Side(style='thin', color="DDDDDD"), bottom=Side(style='thin', color="DDDDDD")
        )
        
        fill_headers = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        font_headers = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        fill_totales = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        colores_destinos = ["D9E2EC", "F0F4F8"]

        ws['A1'] = f"{title_text.upper()}"
        ws['A1'].font = font_titulo
        ws['A1'].alignment = left_aligned

        headers = [
            "Orden de Entrega", "Cliente", "RUT Cliente", "Proveedor", "N° Factura", 
            "Código Interno", "Monto", "IVA", "Total con iva", "Pago Real", 
            "Resp de Pago", "Estado", "Forma de Pago", "Fecha", "Entregado", "Paga Iva"
        ]

        for col_idx, text in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=text)
            cell.font = font_headers
            cell.fill = fill_headers
            cell.alignment = center_aligned
            cell.border = thin_border

        row_num = 4
        ciudad_actual = None
        indice_color = -1
        start_row_data = row_num

        for m in mercancias_list:
            ciudad_item = getattr(m.id_destino, 'nombre_ciudad', 'Sin Ciudad')
            if ciudad_item != ciudad_actual:
                indice_color = (indice_color + 1) % len(colores_destinos)
                relleno_zona = PatternFill(start_color=colores_destinos[indice_color], end_color=colores_destinos[indice_color], fill_type="solid")
                
                ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=16)
                cell = ws.cell(row=row_num, column=1, value=f"{ciudad_item.upper()}")
                cell.font = bold_font
                cell.alignment = left_aligned
                for col in range(1, 17):
                    celda_enc = ws.cell(row=row_num, column=col)
                    celda_enc.border = thin_border
                    celda_enc.fill = relleno_zona
                row_num += 1
                ciudad_actual = ciudad_item

            if is_general:
                rut_cliente_desencriptado = "S/R"
                if m.id_cliente:
                    campo_rut = getattr(m.id_cliente, 'rut_cliente_cifrado', getattr(m.id_cliente, 'rut_cifrado', getattr(m.id_cliente, 'rut_cliente', None)))
                    if campo_rut:
                        try:
                            if str(campo_rut).startswith('g9g') or len(str(campo_rut)) > 30:
                                rut_cliente_desencriptado = fernet.decrypt(campo_rut.encode('utf-8')).decode('utf-8')
                            else:
                                rut_cliente_desencriptado = str(campo_rut)
                        except Exception:
                            rut_cliente_desencriptado = str(campo_rut)
                try:
                    factura_key = int(m.factura) if m.factura else None
                except ValueError:
                    factura_key = None
                doc_contable = documentos_map.get(factura_key) if factura_key else None
                monto_neto = float(doc_contable.subtotal) if doc_contable else float(m.precio_total or 0.0)
                valor_iva = round(monto_neto * 0.19)
                venta_final = monto_neto + valor_iva
                is_facturado = bool(m.factura or doc_contable)
                factura_lbl = "Facturado" if is_facturado else "Sin Facturar"
                is_pagado = (doc_contable and doc_contable.estado == 'Pagado') or (m.estado_cobranza == 'Pagado')
                pago_lbl = "PAGADO" if is_pagado else "NO PAGADO"
                estado_fila = f"{pago_lbl} - {factura_lbl}"
                paga_iva_inicial = "✓" if is_facturado else "✗"
                entregado_inicial = "✓" if is_pagado or estado_fila == "Facturado NO PAGADO" else "✗"
                if doc_contable:
                    forma_pago = 'credito' if doc_contable.condicion_pago != 'Contra_Entrega' else 'transferencia'
                    fecha_fila = doc_contable.fecha_emision.strftime('%d/%m/%Y') if hasattr(doc_contable.fecha_emision, 'strftime') else str(doc_contable.fecha_emision)
                else:
                    forma_pago = 'no facturado'
                    fecha_fila = m.fecha_ingreso.strftime('%d/%m/%Y') if hasattr(m.fecha_ingreso, 'strftime') else str(m.fecha_ingreso)[:10]
                formula_pago_real = f'=IF(P{row_num}="✓", I{row_num}, G{row_num})'
                datos_fila = [
                    m.numero_orden_entrega or 'S/N',
                    getattr(m.id_cliente, 'nombre_cliente', 'S/N'),
                    rut_cliente_desencriptado,
                    getattr(m.id_proveedor, 'nombre_proveedor', 'Sin Proveedor') if m.id_proveedor else 'Sin Proveedor',
                    m.factura or 'S/N',
                    m.codigo_interno or 'S/N',
                    monto_neto, valor_iva, venta_final, formula_pago_real, 
                    'PROVEEDOR' if getattr(m, 'paga_proveedor', False) else 'CLIENTE',
                    estado_fila, forma_pago, fecha_fila, entregado_inicial, paga_iva_inicial
                ]
                if mercancia_to_general_row is not None:
                    mercancia_to_general_row[m.id_mercancia] = row_num
            else:
                gen_row = mercancia_to_general_row.get(m.id_mercancia, 4)
                datos_fila = [
                    f"='Detalle General'!A{gen_row}", f"='Detalle General'!B{gen_row}",
                    f"='Detalle General'!C{gen_row}", f"='Detalle General'!D{gen_row}",
                    f"='Detalle General'!E{gen_row}", f"='Detalle General'!F{gen_row}",
                    f"='Detalle General'!G{gen_row}", f"='Detalle General'!H{gen_row}",
                    f"='Detalle General'!I{gen_row}", f"='Detalle General'!J{gen_row}",
                    f"='Detalle General'!K{gen_row}", f"='Detalle General'!L{gen_row}",
                    f"='Detalle General'!M{gen_row}", f"='Detalle General'!N{gen_row}",
                    f"='Detalle General'!O{gen_row}", f"='Detalle General'!P{gen_row}" 
                ]
            for col_num, dato in enumerate(datos_fila, 1):
                cell = ws.cell(row=row_num, column=col_num, value=dato)
                cell.font = font_regular
                cell.border = thin_border
                
                if col_num in [1, 2, 4]:
                    cell.alignment = left_aligned
                elif col_num in [7, 8, 9, 10]:
                    cell.alignment = right_aligned
                    if is_general or col_num != 10: 
                        cell.number_format = '"$"#,##0'
                else:
                    cell.alignment = center_aligned
            row_num += 1

        end_row_data = row_num - 1
        if is_general:
            dv_checkbox = DataValidation(type="list", formula1='"✓,✗"', allow_blank=True)
            dv_forma_pago = DataValidation(type="list", formula1='"no facturado,efectivo,cheque,credito,transferencia"', allow_blank=True)
            dv_estados_matriz = DataValidation(type="list", formula1='"PAGADO - Facturado,NO PAGADO - Facturado,PAGADO - Sin Facturar,NO PAGADO - Sin Facturar"', allow_blank=True)
            ws.add_data_validation(dv_checkbox)
            ws.add_data_validation(dv_forma_pago)
            ws.add_data_validation(dv_estados_matriz)
            dv_checkbox.add(f"O{start_row_data}:O{end_row_data}")
            dv_checkbox.add(f"P{start_row_data}:P{end_row_data}")
            dv_forma_pago.add(f"M{start_row_data}:M{end_row_data}")
            dv_estados_matriz.add(f"L{start_row_data}:L{end_row_data}")
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
        ws.cell(row=row_num, column=1, value="VALOR DEL VIAJE: ").alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row_num, column=7, value=f"=SUM(G{start_row_data}:G{end_row_data})").number_format = '"$"#,##0'
        ws.cell(row=row_num, column=8, value=f"=SUM(H{start_row_data}:H{end_row_data})").number_format = '"$"#,##0'
        ws.cell(row=row_num, column=9, value=f"=SUM(I{start_row_data}:I{end_row_data})").number_format = '"$"#,##0'
        ws.cell(row=row_num, column=10, value=f"=SUM(J{start_row_data}:J{end_row_data})").number_format = '"$"#,##0'
        for c in range(1, 17):
            celda = ws.cell(row=row_num, column=c)
            celda.font = bold_font
            celda.border = thin_border
            celda.fill = fill_totales
        r_sin_resta = row_num + 1
        r_real = row_num + 2
        r_faltante = row_num + 3
        r_credito = row_num + 4
        r_real_cred = row_num + 5
        r_restante = row_num + 6
        metrics = [
            ("Total sin resta", f"=J{row_num}", "FFFFFF", r_sin_resta, 9),
            ("Total real", f'=SUMIF(L{start_row_data}:L{end_row_data},"PAGADO*",J{start_row_data}:J{end_row_data})', "DCFCE7", r_real, 9), 
            ("Faltante", f"=J{r_sin_resta}-J{r_real}", "FEF08A", r_faltante, 9), 
            ("Total Credito", f'=SUMIF(M{start_row_data}:M{end_row_data},"credito",J{start_row_data}:J{end_row_data})', "DBEAFE", r_credito, 11), 
            ("Real+Credito", f"=J{r_real}+L{r_credito}", "F1F5F9", r_real_cred, 11),
            ("Restante", f"=J{r_sin_resta}-L{r_real_cred}", "FEE2E2", r_restante, 11) 
        ]

        for label, formula, color_hex, target_row, target_col in metrics:
            lbl_c = ws.cell(row=target_row, column=target_col, value=label)
            lbl_c.font = bold_font
            lbl_c.alignment = right_aligned
            lbl_c.border = thin_border
            
            val_c = ws.cell(row=target_row, column=target_col + 1, value=formula)
            val_c.font = bold_font
            val_c.alignment = right_aligned
            val_c.number_format = '"$"#,##0'
            val_c.border = thin_border
            val_c.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        return {
            "total_con_iva": f"='{ws.title}'!J{r_sin_resta}",
            "contado": f"='{ws.title}'!J{r_real}",
            "credito": f"='{ws.title}'!L{r_credito}",
            "falta_cobrar": f"='{ws.title}'!L{r_restante}"
        }

    def retrieve(self, request, *args, **kwargs):
        despacho = self.get_object()
        mercancias = Mercancia.activos.filter(id_despacho=despacho).order_by(
            'id_destino__nombre_ciudad', 'id_cliente__nombre_cliente'  
        ).select_related('id_cliente', 'id_proveedor', 'id_destino')
        facturas_viaje = []
        for m in mercancias:
            if m.factura:
                try: 
                    facturas_viaje.append(int(m.factura))
                except ValueError: 
                    pass
        documentos_map = {
            doc.numero_documento: doc 
            for doc in DocumentoCobro.objects.filter(numero_documento__in=facturas_viaje, activo=True)
        }
        wb = openpyxl.Workbook()
        fernet = Fernet(settings.FIELD_ENCRYPTION_KEY.encode())
        mercancia_to_general_row = {}
        ws_general = wb.active
        ws_general.title = "Detalle General"
        self.write_detailed_sheet(
            ws_general, f"Consolidado General - {despacho.id_ruta or 'S/N'}", 
            mercancias, documentos_map, fernet, is_general=True, 
            mercancia_to_general_row=mercancia_to_general_row
        )
        bucket_zonas = {"IQUIQUE": [], "ANTOFA": [], "COPIAPO": []}
        for m in mercancias:
            zona = self.classify_macro_zone(getattr(m.id_destino, 'nombre_ciudad', ''))
            bucket_zonas[zona].append(m)
        referencias_resumen = {}
        for zona_key, lista_mercancias in bucket_zonas.items():
            if len(lista_mercancias) > 0:
                ws_zona = wb.create_sheet(title=zona_key)
                referencias_resumen[zona_key] = self.write_detailed_sheet(
                    ws_zona, f"Lotes Logísticos Bodega {zona_key}", 
                    lista_mercancias, documentos_map, fernet, is_general=False, 
                    mercancia_to_general_row=mercancia_to_general_row
                )
        ws_resumen = wb.create_sheet(title="Resumen Consolidado")
        ws_resumen.views.sheetView[0].showGridLines = True
        fill_res_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") 
        fill_alert_red = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") 
        fill_blue_total = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        border_box = Border(left=Side(style='thin', color="000000"), right=Side(style='thin', color="000000"), top=Side(style='thin', color="000000"), bottom=Side(style='thin', color="000000"))
        font_res_bold = Font(name="Arial", size=10, bold=True)
        res_headers = ["TOTALES", "CREDITO", "AL CONTADO", "SUMA", "FALTA COBRAR", "COBRADO"]
        for c_idx, h_text in enumerate(res_headers, 1):
            cell = ws_resumen.cell(row=2, column=c_idx, value=h_text)
            cell.font = font_res_bold
            cell.fill = fill_res_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_box
        filas_zonas = ["IQUIQUE", "ANTOFA", "COPIAPO"]
        current_r = 3
        for zona in filas_zonas:
            ws_resumen.cell(row=current_r, column=1, value=zona).font = font_res_bold
            ws_resumen.cell(row=current_r, column=1).alignment = Alignment(horizontal="left")
            ws_resumen.cell(row=current_r, column=1).border = border_box
            ref = referencias_resumen.get(zona, {"credito": "0", "contado": "0", "falta_cobrar": "0", "total_con_iva": "0"})
            ws_resumen.cell(row=current_r, column=2, value=ref["credito"]).number_format = '"$"#,##0'
            ws_resumen.cell(row=current_r, column=3, value=ref["contado"]).number_format = '"$"#,##0'
            ws_resumen.cell(row=current_r, column=4, value=f"=B{current_r}+C{current_r}").number_format = '"$"#,##0' 
            cell_falta = ws_resumen.cell(row=current_r, column=5, value=ref["falta_cobrar"])
            cell_falta.number_format = '"$"#,##0'
            cell_falta.fill = fill_alert_red
            ws_resumen.cell(row=current_r, column=6, value="").border = border_box 
            for c in range(2, 7):
                ws_resumen.cell(row=current_r, column=c).border = border_box
                ws_resumen.cell(row=current_r, column=c).font = font_res_bold
                ws_resumen.cell(row=current_r, column=c).alignment = Alignment(horizontal="right")
            current_r += 1
        ws_resumen.cell(row=current_r, column=1, value="TOTALES").font = font_res_bold
        ws_resumen.cell(row=current_r, column=1).border = border_box
        ws_resumen.cell(row=current_r, column=2, value="=SUM(B3:B5)").number_format = '"$"#,##0'
        ws_resumen.cell(row=current_r, column=3, value="=SUM(C3:C5)").number_format = '"$"#,##0'
        ws_resumen.cell(row=current_r, column=4, value="=SUM(D3:D5)").number_format = '"$"#,##0'
        ws_resumen.cell(row=current_r, column=5, value="=SUM(E3:E5)").number_format = '"$"#,##0'
        cell_final_cobrado = ws_resumen.cell(row=current_r, column=6, value="=SUM(C3:C5)")
        cell_final_cobrado.number_format = '"$"#,##0'
        cell_final_cobrado.fill = fill_blue_total
        for c in range(1, 7):
            cell = ws_resumen.cell(row=current_r, column=c)
            cell.font = font_res_bold
            cell.border = border_box
            if c > 1: 
                cell.alignment = Alignment(horizontal="right")
        for col in ws_resumen.columns:
            ws_resumen.column_dimensions[get_column_letter(col[0].column)].width = 18
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Control_Facturacion_Despacho_{despacho.id_despacho}.xlsx"'
        wb.save(response)
        
        return response
    

class FinanzasMercanciasExcelAPIView(generics.ListAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, *args, **kwargs):
        empresa_id = request.user.perfil.empresa_id
        
        start_row = int(request.query_params.get('startRow', 0))
        end_row = int(request.query_params.get('endRow', 100))
        
        cliente_filtro = request.query_params.get('cliente_nombre', None)
        ruta_filtro = request.query_params.get('codigo_ruta', None)
        estado_filtro = request.query_params.get('estado_cobranza', None)

        queryset = Mercancia.objects.filter(
            empresa_id=empresa_id,
            activo=True
        )

        if estado_filtro:
            queryset = queryset.filter(estado_cobranza=estado_filtro)
        else:
            queryset = queryset.filter(estado_cobranza__in=['Pendiente', 'En_Proceso'])

        if cliente_filtro:
            queryset = queryset.filter(id_cliente__nombre_cliente__icontains=cliente_filtro)
            
        if ruta_filtro:
            queryset = queryset.filter(id_despacho__id_ruta__codigo_ruta__icontains=ruta_filtro)

        total_records = queryset.count()

        records = queryset.select_related(
            'id_cliente', 'id_despacho__id_ruta', 'id_destino'
        ).prefetch_related(
            'documentos_cobro_asociados'
        )[start_row:end_row]

        serializer = MercanciaPendienteCobroSerializer(records, many=True)

        return Response({
            "rows": serializer.data,
            "total": total_records
        }, status=status.HTTP_200_OK)